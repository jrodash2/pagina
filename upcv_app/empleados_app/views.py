import json
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from uuid import uuid4

from PIL import Image, ImageColor, ImageDraw, ImageFont, ImageOps
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.db import OperationalError, ProgrammingError, transaction
from django.db.models import Case, Count, IntegerField, Q, Sum, When
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import NoReverseMatch, reverse
from django.views.decorators.http import require_GET, require_POST

from .forms import (
    CarreraForm,
    CargaMasivaExcelForm,
    ConfiguracionGeneralForm,
    EmpleadoEditForm,
    EmpleadoForm,
    EstablecimientoForm,
    GradoForm,
    MatriculaForm,
    MatriculaMasivaForm,
    ObservacionAlumnoForm,
    UsuarioCreateForm,
    UsuarioUpdateForm,
)
from .gafete_utils import (
    canvas_for_orientation,
    is_item_allowed_in_face,
    is_item_visible_in_face,
    normalizar_layout_gafete,
    obtener_layout_cara,
    orientation_for_establecimiento,
    resolve_gafete_dimensions,
    serializar_layout_frente_reverso,
)
from .models import Asistencia, AsistenciaDetalle, CicloEscolar, Curso, CursoDocente, DEFAULT_GAFETE_LAYOUT, Carrera, ConfiguracionGeneral, Empleado, Establecimiento, Grado, Matricula, ObservacionAlumno, Perfil
from .permissions import (
    es_admin_total,
    es_docente,
    es_gestor,
    filtrar_por_establecimiento_usuario,
    obtener_establecimiento_usuario,
    puede_acceder_backoffice,
    puede_administrar_configuracion,
    puede_operar_establecimiento,
    usuario_puede_ver_establecimiento,
)


def _can_manage_design(user):
    return es_admin_total(user)


def _is_docente(user):
    return es_docente(user)


def _can_access_backoffice(user):
    return puede_acceder_backoffice(user)


def _can_access_alumnos(user):
    return bool(user and user.is_authenticated and (puede_acceder_backoffice(user) or es_docente(user)))


def _can_access_admin_config(user):
    return puede_administrar_configuracion(user)


def _can_manage_establecimiento(user):
    return puede_operar_establecimiento(user)


def _deny_if_not_allowed_establecimiento(request, establecimiento_id):
    if not usuario_puede_ver_establecimiento(request.user, establecimiento_id):
        messages.error(request, "No tiene permisos para acceder a ese establecimiento.")
        return redirect("empleados:dahsboard")
    return None


def _forbid_gafetes_for_gestor(request):
    if es_gestor(request.user):
        return HttpResponseForbidden("No tiene permiso para acceder a gafetes.")
    return None


def _docente_alumnos_qs(user):
    grados_asignados = (
        CursoDocente.objects.filter(
            docente=user,
            activo=True,
            curso__activo=True,
        )
        .values_list("curso__grado_id", flat=True)
        .distinct()
    )
    return (
        Empleado.objects.filter(
            activo=True,
            matriculas__estado="activo",
            matriculas__grado_id__in=grados_asignados,
        )
        .select_related("establecimiento", "grado")
        .distinct()
        .order_by("-created_at")
    )


def _sanitize_face_items(items, enabled_fields, canvas_width, canvas_height, allow_empty=False):
    base_keys = {"photo", "nombres", "apellidos", "codigo_alumno", "grado", "grado_descripcion", "sitio_web", "telefono", "cui", "establecimiento", "texto_libre_1", "texto_libre_2", "texto_libre_3", "image"}
    allowed_align = {"left", "center", "right"}
    allowed_weight = {"400", "700"}
    allowed_fit = {"contain", "cover"}

    result_items = {}
    valid_enabled = [
        field for field in (enabled_fields or [])
        if field in base_keys or str(field).startswith("texto_libre_") or str(field).startswith("image")
    ]

    for key, cfg in (items or {}).items():
        is_dynamic_text = str(key).startswith("texto_libre_")
        is_dynamic_image = str(key).startswith("image")
        if (key not in base_keys and not is_dynamic_text and not is_dynamic_image) or not isinstance(cfg, dict):
            continue

        if key == "photo":
            border_color = (cfg.get("border_color") or "#ffffff").strip()
            if not re.fullmatch(r"#[0-9a-fA-F]{6}", border_color):
                raise ValueError("Color de borde inválido para photo")
            shape = (cfg.get("shape") or "rounded").strip().lower()
            if shape not in {"rounded", "circle"}:
                raise ValueError("Forma inválida para photo")
            result_items[key] = {
                "x": int(cfg.get("x") or 0),
                "y": int(cfg.get("y") or 0),
                "w": max(40, min(canvas_width, int(cfg.get("w") or 250))),
                "h": max(40, min(canvas_height, int(cfg.get("h") or 350))),
                "shape": shape,
                "radius": max(0, min(200, int(cfg.get("radius") or 20))),
                "border": bool(cfg.get("border", True)),
                "border_width": max(0, min(20, int(cfg.get("border_width") or 4))),
                "border_color": border_color,
                "visible": bool(cfg.get("visible", True)),
            }
            continue

        if key == "image" or is_dynamic_image:
            fit = str(cfg.get("object_fit") or "contain").lower()
            result_items[key] = {
                "x": int(cfg.get("x") or 0),
                "y": int(cfg.get("y") or 0),
                "w": max(40, min(canvas_width, int(cfg.get("w") or 220))),
                "h": max(40, min(canvas_height, int(cfg.get("h") or 220))),
                "src": str(cfg.get("src") or ""),
                "object_fit": fit if fit in allowed_fit else "contain",
                "visible": bool(cfg.get("visible", False)),
            }
            continue

        color = (cfg.get("color") or "#111111").strip()
        if not re.fullmatch(r"#[0-9a-fA-F]{6}", color):
            raise ValueError(f"Color inválido para {key}")
        align = (cfg.get("align") or "left").strip().lower()
        if align not in allowed_align:
            align = "left"
        weight = str(cfg.get("font_weight") or "400")
        if weight not in allowed_weight:
            weight = "400"
        item = {
            "x": int(cfg.get("x") or 0),
            "y": int(cfg.get("y") or 0),
            "w": max(40, min(canvas_width, int(cfg.get("w") or 280))),
            "h": max(30, min(canvas_height, int(cfg.get("h") or 70))),
            "font_size": max(10, min(120, int(cfg.get("font_size") or 24))),
            "font_weight": weight,
            "color": color,
            "align": align,
            "visible": bool(cfg.get("visible", True)),
        }
        if "w" in cfg:
            item["w"] = max(40, min(canvas_width, int(cfg.get("w") or 280)))
        if "h" in cfg:
            item["h"] = max(30, min(canvas_height, int(cfg.get("h") or 70)))
        if str(key).startswith("texto_libre_"):
            item["text"] = str(cfg.get("text") or "")
        result_items[key] = item

    if not result_items and not allow_empty:
        raise ValueError("No se recibieron items válidos")
    return result_items, valid_enabled


def _validate_layout_payload(payload, forced_orientation=None):
    if not isinstance(payload, dict):
        raise ValueError("Formato inválido")

    layout = payload.get("layout", payload)
    if not isinstance(layout, dict):
        raise ValueError("Layout inválido")

    normalized = normalizar_layout_gafete(layout, orientation=forced_orientation or "H")
    canvas = normalized.get("canvas") or {}
    orientation = str((forced_orientation or canvas.get("orientation") or "H")).upper()
    if orientation not in {"H", "V"}:
        orientation = "H"
    canvas_width, canvas_height = canvas_for_orientation(orientation)

    out = {
        "canvas": {"width": canvas_width, "height": canvas_height, "orientation": orientation},
        "front": {},
        "back": {},
    }

    for face in ("front", "back"):
        face_layout = normalized.get(face, {}) if isinstance(normalized, dict) else {}
        face_items, enabled = _sanitize_face_items(
            face_layout.get("items"),
            face_layout.get("enabled_fields"),
            canvas_width,
            canvas_height,
            allow_empty=(face == "back"),
        )
        if face == "back":
            enabled = [field for field in enabled if is_item_allowed_in_face("back", field)]
        out[face] = {
            "background_image": str(face_layout.get("background_image") or ""),
            "enabled_fields": enabled,
            "items": face_items,
        }

    return out


def _canvas_dimensions(establecimiento, orientation=None):
    orient = orientation or orientation_for_establecimiento(establecimiento)
    return canvas_for_orientation(orient)


def home(request):
    return render(request, "empleados/login.html")

def home(request):
    return render(request, "empleados/login.html")

def signin(request):
    if request.method == "GET":
        return render(request, "empleados/login.html", {"form": AuthenticationForm})

    user = authenticate(request, username=request.POST["username"], password=request.POST["password"])
    if user is None:
        return render(request, "empleados/login.html", {"form": AuthenticationForm, "error": "Usuario o Password es Incorrecto"})

    auth_login(request, user)

    if user.groups.filter(name="Docente").exists():
        redirect_name = "empleados:dashboard_docente"
    elif user.groups.filter(name="Administrador").exists():
        redirect_name = "dashboard"
    elif user.groups.filter(name="Gestor").exists():
        redirect_name = "dashboard_gestor"
    elif user.groups.filter(name="Departamento").exists():
        redirect_name = "dashboard_departamento"
    else:
        redirect_name = "dashboard"

    for target in (redirect_name, "empleados:dahsboard"):
        try:
            return redirect(target)
        except NoReverseMatch:
            continue

    raise


def signout(request):
    logout(request)
    return redirect("empleados:signin")


@login_required
@user_passes_test(_can_access_admin_config)
def usuarios_list(request):
    usuarios = User.objects.prefetch_related("groups").all().order_by("username")
    usuarios_rows = []

    for usuario in usuarios:
        foto_url = ""
        try:
            perfil, _ = Perfil.objects.get_or_create(user=usuario)
            foto_url = perfil.foto.url if perfil.foto else ""
        except (ProgrammingError, OperationalError):
            foto_url = ""

        usuarios_rows.append({"usuario": usuario, "foto_url": foto_url})

    return render(request, "empleados/usuarios_list.html", {"usuarios_rows": usuarios_rows})


@login_required
@user_passes_test(_can_access_admin_config)
def usuarios_create(request):
    form = UsuarioCreateForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        usuario = form.save()
        perfil, _ = Perfil.objects.get_or_create(user=usuario)
        foto = form.cleaned_data.get("foto")
        if foto:
            perfil.foto = foto
            perfil.save()
        messages.success(request, "Usuario creado correctamente.")
        return redirect("empleados:usuarios_list")
    return render(request, "empleados/usuarios_form.html", {"form": form, "titulo": "Nuevo Usuario", "perfil_foto_url": ""})


@login_required
@user_passes_test(_can_access_admin_config)
def usuarios_update(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    form = UsuarioUpdateForm(request.POST or None, request.FILES or None, instance=usuario)
    if request.method == "POST" and form.is_valid():
        usuario = form.save()
        perfil, _ = Perfil.objects.get_or_create(user=usuario)
        foto = form.cleaned_data.get("foto")
        if foto:
            perfil.foto = foto
            perfil.save()
        messages.success(request, "Usuario actualizado correctamente.")
        return redirect("empleados:usuarios_list")
    perfil = None
    try:
        perfil, _ = Perfil.objects.get_or_create(user=usuario)
    except (ProgrammingError, OperationalError):
        perfil = None
    perfil_foto_url = perfil.foto.url if perfil and perfil.foto else ""
    return render(request, "empleados/usuarios_form.html", {"form": form, "titulo": "Editar Usuario", "usuario": usuario, "perfil_foto_url": perfil_foto_url})


@login_required
def dahsboard(request):
    if _is_docente(request.user):
        return redirect("empleados:dashboard_docente")

    establecimiento_param = (request.GET.get("establecimiento") or "").strip()
    establecimientos_qs = filtrar_por_establecimiento_usuario(Establecimiento.objects.order_by("nombre"), request.user, "id")

    selected_establecimiento = None
    if establecimiento_param and establecimiento_param != "all":
        selected_establecimiento = establecimientos_qs.filter(pk=establecimiento_param).first()

    if es_gestor(request.user):
        selected_establecimiento = selected_establecimiento or establecimientos_qs.first()

    alumnos_base = Matricula.objects.select_related("grado__carrera__ciclo_escolar__establecimiento")
    cursos_docente_base = CursoDocente.objects.select_related("curso__grado__carrera__ciclo_escolar__establecimiento", "docente")
    cursos_base = Curso.objects.select_related("grado__carrera__ciclo_escolar__establecimiento")
    asistencia_base = Asistencia.objects.select_related("curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento")
    detalles_base = AsistenciaDetalle.objects.select_related("asistencia__curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento")

    if selected_establecimiento:
        alumnos_base = alumnos_base.filter(grado__carrera__ciclo_escolar__establecimiento=selected_establecimiento)
        cursos_docente_base = cursos_docente_base.filter(curso__grado__carrera__ciclo_escolar__establecimiento=selected_establecimiento)
        cursos_base = cursos_base.filter(grado__carrera__ciclo_escolar__establecimiento=selected_establecimiento)
        asistencia_base = asistencia_base.filter(curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento=selected_establecimiento)
        detalles_base = detalles_base.filter(asistencia__curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento=selected_establecimiento)

    alumnos_por_grado_qs = (
        alumnos_base.values("grado__nombre")
        .annotate(total=Count("alumno", distinct=True))
        .order_by("-total", "grado__nombre")
    )
    alumnos_por_grado_labels = [row["grado__nombre"] or "Sin grado" for row in alumnos_por_grado_qs]
    alumnos_por_grado_series = [row["total"] for row in alumnos_por_grado_qs]

    alumnos_por_carrera_qs = (
        alumnos_base.values("grado__carrera__nombre")
        .annotate(total=Count("alumno", distinct=True))
        .order_by("-total", "grado__carrera__nombre")
    )
    alumnos_por_carrera_labels = [row["grado__carrera__nombre"] or "Sin carrera" for row in alumnos_por_carrera_qs]
    alumnos_por_carrera_series = [row["total"] for row in alumnos_por_carrera_qs]

    cursos_por_docente_qs = (
        cursos_docente_base.filter(activo=True, curso__activo=True)
        .values("docente__first_name", "docente__last_name", "docente__username")
        .annotate(total=Count("curso", distinct=True))
        .order_by("-total", "docente__username")[:12]
    )
    cursos_por_docente_labels, cursos_por_docente_series = [], []
    for row in cursos_por_docente_qs:
        nombre = f"{(row['docente__first_name'] or '').strip()} {(row['docente__last_name'] or '').strip()}".strip() or row["docente__username"]
        cursos_por_docente_labels.append(nombre)
        cursos_por_docente_series.append(row["total"])

    presentes_total = detalles_base.filter(presente=True).count()
    ausentes_total = detalles_base.filter(presente=False).count()
    asistencia_resumen_series = [presentes_total, ausentes_total]

    tendencia_qs = (
        asistencia_base.values("fecha")
        .annotate(
            presentes=Count("detalles", filter=Q(detalles__presente=True)),
            ausentes=Count("detalles", filter=Q(detalles__presente=False)),
        )
        .order_by("-fecha")[:14]
    )
    tendencia = list(reversed(list(tendencia_qs)))
    asistencia_tendencia_labels = [row["fecha"].strftime("%d/%m") for row in tendencia]
    asistencia_tendencia_presentes = [row["presentes"] for row in tendencia]
    asistencia_tendencia_ausentes = [row["ausentes"] for row in tendencia]

    establecimientos_ids = list(establecimientos_qs.values_list("id", flat=True))

    cursos_por_establecimiento = {
        row["grado__carrera__ciclo_escolar__establecimiento_id"]: row["total"]
        for row in Curso.objects.filter(activo=True, grado__carrera__ciclo_escolar__establecimiento_id__in=establecimientos_ids)
        .values("grado__carrera__ciclo_escolar__establecimiento_id")
        .annotate(total=Count("id", distinct=True))
    }
    alumnos_por_establecimiento = {
        row["grado__carrera__ciclo_escolar__establecimiento_id"]: row["total"]
        for row in Matricula.objects.filter(grado__carrera__ciclo_escolar__establecimiento_id__in=establecimientos_ids)
        .values("grado__carrera__ciclo_escolar__establecimiento_id")
        .annotate(total=Count("alumno_id", distinct=True))
    }
    asistencias_por_establecimiento = {
        row["curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento_id"]: row["total"]
        for row in Asistencia.objects.filter(curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento_id__in=establecimientos_ids)
        .values("curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento_id")
        .annotate(total=Count("id"))
    }

    establecimientos_resumen = []
    for est in establecimientos_qs:
        est_id = est.id
        establecimientos_resumen.append({
            "id": est_id,
            "nombre": est.nombre or "Sin establecimiento",
            "cursos": cursos_por_establecimiento.get(est_id, 0),
            "alumnos": alumnos_por_establecimiento.get(est_id, 0),
            "asistencias": asistencias_por_establecimiento.get(est_id, 0),
        })

    establecimientos_destacados = []
    if not selected_establecimiento:
        establecimientos_destacados = sorted(
            establecimientos_resumen,
            key=lambda x: (x["alumnos"], x["cursos"], x["asistencias"]),
            reverse=True,
        )[:6]
    titulo_dashboard = (
        f"Dashboard de {selected_establecimiento.nombre}" if selected_establecimiento else "Dashboard global"
    )
    subtitulo_dashboard = (
        "Resumen del establecimiento seleccionado." if selected_establecimiento else "Resumen general de todos los establecimientos."
    )

    ciclo_stats = {
        "nombre": selected_establecimiento.nombre if selected_establecimiento else "Todos los establecimientos",
        "establecimiento": selected_establecimiento.nombre if selected_establecimiento else "Global",
        "cursos_total": cursos_base.filter(activo=True).count(),
        "alumnos_total": alumnos_base.values("alumno_id").distinct().count(),
        "docentes_total": cursos_docente_base.filter(activo=True).values("docente_id").distinct().count(),
        "asistencias_total": asistencia_base.count(),
        "establecimientos_total": establecimientos_qs.count() if es_admin_total(request.user) else 1,
    }

    context = {
        "titulo_dashboard": titulo_dashboard,
        "subtitulo_dashboard": subtitulo_dashboard,
        "selected_establecimiento": selected_establecimiento,
        "establecimientos": establecimientos_qs,
        "ciclo_activo": True,
        "ciclo_stats": ciclo_stats,
        "establecimientos_resumen": establecimientos_resumen,
        "establecimientos_destacados": establecimientos_destacados,
        "alumnos_por_grado_labels": json.dumps(alumnos_por_grado_labels),
        "alumnos_por_grado_series": json.dumps(alumnos_por_grado_series),
        "alumnos_por_carrera_labels": json.dumps(alumnos_por_carrera_labels),
        "alumnos_por_carrera_series": json.dumps(alumnos_por_carrera_series),
        "cursos_por_docente_labels": json.dumps(cursos_por_docente_labels),
        "cursos_por_docente_series": json.dumps(cursos_por_docente_series),
        "asistencia_resumen_series": json.dumps(asistencia_resumen_series),
        "asistencia_tendencia_labels": json.dumps(asistencia_tendencia_labels),
        "asistencia_tendencia_presentes": json.dumps(asistencia_tendencia_presentes),
        "asistencia_tendencia_ausentes": json.dumps(asistencia_tendencia_ausentes),
    }
    return render(request, "empleados/dahsboard.html", context)


@login_required
def dashboard_establecimiento(request, establecimiento_id):
    if _is_docente(request.user):
        return redirect("empleados:dashboard_docente")
    return redirect(f"{reverse('empleados:dahsboard')}?establecimiento={establecimiento_id}")


@login_required
@user_passes_test(_can_access_admin_config)
def configuracion_general(request):
    configuracion, _ = ConfiguracionGeneral.objects.get_or_create(id=1)
    form = ConfiguracionGeneralForm(request.POST or None, request.FILES or None, instance=configuracion)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Configuración actualizada.")
        return redirect("empleados:configuracion_general")
    return render(request, "empleados/configuracion_general.html", {"form": form, "configuracion": configuracion})


@login_required
@user_passes_test(_can_access_backoffice)
def crear_empleado(request):
    form = EmpleadoForm(request.POST or None, request.FILES or None)
    establecimiento_usuario = obtener_establecimiento_usuario(request.user)
    if establecimiento_usuario and "grado" in form.fields:
        form.fields["grado"].queryset = form.fields["grado"].queryset.filter(carrera__ciclo_escolar__establecimiento=establecimiento_usuario)
    if request.method == "POST" and form.is_valid():
        empleado = form.save(commit=False)
        empleado.user = request.user
        if establecimiento_usuario:
            empleado.establecimiento = establecimiento_usuario
        empleado.save()
        messages.success(request, "Alumno creado correctamente.")
        return redirect("empleados:empleado_lista")
    return render(request, "empleados/crear_empleado.html", {"form": form, "grados": Grado.objects.all()})


@login_required
@user_passes_test(_can_access_backoffice)
def editar_empleado(request, e_id):
    empleado = get_object_or_404(Empleado, pk=e_id)
    if empleado.establecimiento_id and not _is_docente(request.user):
        denied = _deny_if_not_allowed_establecimiento(request, empleado.establecimiento_id)
        if denied:
            return denied
    form = EmpleadoEditForm(request.POST or None, request.FILES or None, instance=empleado)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Alumno actualizado correctamente.")
        return redirect("empleados:empleado_lista")
    return render(request, "empleados/editar_empleado.html", {"form": form, "grados": Grado.objects.all(), "empleado": empleado})


@login_required
@user_passes_test(_can_access_alumnos)
def lista_empleados(request):
    if _is_docente(request.user):
        empleados = _docente_alumnos_qs(request.user)
    else:
        empleados = Empleado.objects.all().order_by("-created_at")
        empleados = filtrar_por_establecimiento_usuario(empleados, request.user, "establecimiento_id")
    return render(request, "empleados/lista_empleados.html", {"empleados": empleados})


@login_required
@user_passes_test(_can_access_backoffice)
def credencial_empleados(request):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    empleados = Empleado.objects.all()
    empleados = filtrar_por_establecimiento_usuario(empleados, request.user, "establecimiento_id")
    return render(request, "empleados/credencial_empleados.html", {"empleados": empleados})


@login_required
@user_passes_test(_can_access_alumnos)
def empleado_detalle(request, id):
    if not _is_docente(request.user):
        forbidden = _forbid_gafetes_for_gestor(request)
        if forbidden:
            return forbidden
    empleado = get_object_or_404(Empleado, id=id)
    if _is_docente(request.user) and not _docente_alumnos_qs(request.user).filter(id=empleado.id).exists():
        messages.error(request, "No tiene permisos para ver este alumno.")
        return redirect("empleados:empleado_lista")
    if empleado.establecimiento_id and not _is_docente(request.user):
        denied = _deny_if_not_allowed_establecimiento(request, empleado.establecimiento_id)
        if denied:
            return denied
    configuracion = ConfiguracionGeneral.objects.first()
    matricula_activa = empleado.matriculas.filter(estado="activo").select_related("grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento").first()
    observacion_form = ObservacionAlumnoForm(request.POST or None)
    if request.method == "POST" and request.POST.get("action") == "crear_observacion":
        if observacion_form.is_valid():
            observacion = observacion_form.save(commit=False)
            observacion.alumno = empleado
            observacion.creado_por = request.user
            observacion.save()
            messages.success(request, "Observación registrada correctamente.")
            return redirect("empleados:empleado_detalle", id=empleado.id)
        messages.error(request, "No fue posible guardar la observación. Verifique los datos.")
    observaciones = empleado.observaciones.select_related("creado_por").all()[:30]

    asistencia_resumen = AsistenciaDetalle.objects.filter(alumno=empleado).aggregate(
        total=Count("id"),
        asistencias=Sum(Case(When(presente=True, then=1), default=0, output_field=IntegerField())),
        inasistencias=Sum(Case(When(presente=False, then=1), default=0, output_field=IntegerField())),
    )
    total_registros = asistencia_resumen["total"] or 0
    presentes = asistencia_resumen["asistencias"] or 0
    ausentes = asistencia_resumen["inasistencias"] or 0
    porcentaje_asistencia = round((presentes / total_registros * 100), 2) if total_registros else 0

    establecimiento = None
    grado_gafete = None
    if matricula_activa and matricula_activa.grado:
        grado_gafete = matricula_activa.grado
        if matricula_activa.grado.carrera:
            establecimiento = matricula_activa.grado.carrera.ciclo_escolar.establecimiento

    orientation = orientation_for_establecimiento(establecimiento)
    layout = normalizar_layout_gafete(establecimiento.get_layout() if establecimiento else {}, orientation=orientation)
    canvas_width, canvas_height = canvas_for_orientation(orientation)
    layout["canvas"] = {"width": canvas_width, "height": canvas_height, "orientation": orientation}
    return render(
        request,
        "empleados/empleado_detalle.html",
        {
            "empleado": empleado,
            "configuracion": configuracion,
            "is_editor": True,
            "establecimiento": establecimiento,
            "layout": layout,
            "grado_gafete": grado_gafete,
            "matricula_activa": matricula_activa,
            "canvas_width": canvas_width,
            "canvas_height": canvas_height,
            "gafete_w": canvas_width,
            "gafete_h": canvas_height,
            "orientacion": orientation,
            "observacion_form": observacion_form,
            "observaciones": observaciones,
            "asistencia_resumen": {
                "total": total_registros,
                "presentes": presentes,
                "ausentes": ausentes,
                "porcentaje": porcentaje_asistencia,
            },
        },
    )


@login_required
@user_passes_test(_can_access_alumnos)
def empleado_boleta_asistencia(request, id):
    empleado = get_object_or_404(Empleado, id=id)
    if _is_docente(request.user) and not _docente_alumnos_qs(request.user).filter(id=empleado.id).exists():
        messages.error(request, "No tiene permisos para ver este alumno.")
        return redirect("empleados:empleado_lista")
    if empleado.establecimiento_id and not _is_docente(request.user):
        denied = _deny_if_not_allowed_establecimiento(request, empleado.establecimiento_id)
        if denied:
            return denied

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    detalles_qs = (
        AsistenciaDetalle.objects.filter(alumno=empleado)
        .select_related("asistencia__curso_docente__curso", "asistencia__curso_docente__curso__grado")
        .order_by("-asistencia__fecha", "-id")
    )
    if fecha_inicio:
        detalles_qs = detalles_qs.filter(asistencia__fecha__gte=fecha_inicio)
    if fecha_fin:
        detalles_qs = detalles_qs.filter(asistencia__fecha__lte=fecha_fin)

    resumen = detalles_qs.aggregate(
        total=Count("id"),
        asistencias=Sum(Case(When(presente=True, then=1), default=0, output_field=IntegerField())),
        inasistencias=Sum(Case(When(presente=False, then=1), default=0, output_field=IntegerField())),
    )
    total_registros = resumen["total"] or 0
    asistencias = resumen["asistencias"] or 0
    inasistencias = resumen["inasistencias"] or 0
    porcentaje = round((asistencias / total_registros * 100), 2) if total_registros else 0

    if request.GET.get("formato") == "xlsx":
        try:
            from openpyxl import Workbook
            from .aulapro.excel import autosize_columns, style_table_header, style_table_row, style_title, workbook_to_response
        except ImportError:
            messages.warning(request, "No fue posible generar Excel en este entorno. Se muestra la boleta en pantalla.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Boleta Asistencia"

            style_title(ws, 1, f"Boleta de asistencia · {empleado}", max_col=6)
            style_table_row(ws, 3, ("Desde", fecha_inicio or "Inicio", "Hasta", fecha_fin or "Hoy", "", ""))
            style_table_row(ws, 4, ("Total registros", total_registros, "Asistencias", asistencias, "Inasistencias", inasistencias))
            style_table_row(ws, 5, ("Porcentaje", f"{porcentaje:.2f}%", "", "", "", ""))

            style_table_header(ws, 7, ("Fecha", "Curso", "Grado", "Estado", "Periodo", "Docente"))
            current_row = 8
            for detalle in detalles_qs:
                asistencia = detalle.asistencia
                curso_docente = asistencia.curso_docente
                style_table_row(
                    ws,
                    current_row,
                    (
                        asistencia.fecha.strftime("%d/%m/%Y"),
                        curso_docente.curso.nombre,
                        curso_docente.curso.grado.nombre,
                        "Presente" if detalle.presente else "Ausente",
                        asistencia.periodo.nombre if asistencia.periodo_id else "-",
                        curso_docente.docente.get_full_name() or curso_docente.docente.username,
                    ),
                )
                current_row += 1
            autosize_columns(ws)
            return workbook_to_response(wb, f"boleta_asistencia_{empleado.id}")

    return render(
        request,
        "empleados/boleta_asistencia.html",
        {
            "empleado": empleado,
            "detalles": detalles_qs[:300],
            "resumen": {
                "total": total_registros,
                "asistencias": asistencias,
                "inasistencias": inasistencias,
                "porcentaje": porcentaje,
            },
            "fecha_inicio": fecha_inicio or "",
            "fecha_fin": fecha_fin or "",
        },
    )


@login_required
@user_passes_test(_can_access_backoffice)
def lista_establecimientos(request):
    establecimientos = Establecimiento.objects.all()
    establecimientos = filtrar_por_establecimiento_usuario(establecimientos, request.user, "id")
    return render(request, "empleados/establecimiento_lista.html", {"establecimientos": establecimientos})


@login_required
@user_passes_test(_can_access_admin_config)
def crear_establecimiento(request):
    form = EstablecimientoForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Establecimiento creado.")
        return redirect("empleados:establecimiento_lista")
    return render(request, "empleados/establecimiento_form.html", {"form": form, "titulo": "Crear establecimiento"})


@login_required
@user_passes_test(_can_access_admin_config)
def editar_establecimiento(request, pk):
    establecimiento = get_object_or_404(Establecimiento, pk=pk)
    form = EstablecimientoForm(request.POST or None, request.FILES or None, instance=establecimiento)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Establecimiento actualizado.")
        return redirect("empleados:establecimiento_lista")
    return render(request, "empleados/establecimiento_form.html", {"form": form, "titulo": "Editar establecimiento", "establecimiento": establecimiento})


@login_required
@user_passes_test(_can_access_backoffice)
def lista_carreras(request):
    carreras = Carrera.objects.select_related("ciclo_escolar", "ciclo_escolar__establecimiento")
    carreras = filtrar_por_establecimiento_usuario(carreras, request.user, "ciclo_escolar__establecimiento_id")
    return render(request, "empleados/carrera_lista.html", {"carreras": carreras})


@login_required
@user_passes_test(_can_access_backoffice)
def crear_carrera(request):
    ciclo_id = request.GET.get("ciclo_escolar") or request.POST.get("ciclo_escolar")
    initial = {"ciclo_escolar": ciclo_id} if ciclo_id else None
    form = CarreraForm(request.POST or None, initial=initial)
    if request.method == "POST" and form.is_valid():
        carrera = form.save(commit=False)
        if ciclo_id:
            carrera.ciclo_escolar_id = int(ciclo_id)
        carrera.save()
        messages.success(request, "Carrera creada.")
        return redirect("empleados:ciclo_detail", est_id=carrera.ciclo_escolar.establecimiento_id, ciclo_id=carrera.ciclo_escolar_id)
    return render(request, "empleados/carrera_form.html", {"form": form, "titulo": "Crear carrera"})


@login_required
@user_passes_test(_can_access_backoffice)
def editar_carrera(request, pk):
    carrera = get_object_or_404(Carrera, pk=pk)
    form = CarreraForm(request.POST or None, instance=carrera)
    if request.method == "POST" and form.is_valid():
        carrera = form.save(commit=False)
        carrera.ciclo_escolar = carrera.ciclo_escolar
        carrera.save()
        messages.success(request, "Carrera actualizada.")
        return redirect("empleados:ciclo_detail", est_id=carrera.ciclo_escolar.establecimiento_id, ciclo_id=carrera.ciclo_escolar_id)
    return render(request, "empleados/carrera_form.html", {"form": form, "titulo": "Editar carrera"})


@login_required
@user_passes_test(_can_access_backoffice)
def lista_grados(request):
    grados = Grado.objects.select_related("carrera", "carrera__ciclo_escolar__establecimiento")
    grados = filtrar_por_establecimiento_usuario(grados, request.user, "carrera__ciclo_escolar__establecimiento_id")
    return render(request, "empleados/grado_lista.html", {"grados": grados})


@login_required
@user_passes_test(_can_access_backoffice)
def crear_grado(request):
    form = GradoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        grado = form.save(commit=False)
        carrera_id = request.POST.get("carrera")
        if carrera_id:
            grado.carrera_id = int(carrera_id)
        grado.save()
        messages.success(request, "Grado creado.")
        return redirect("empleados:grado_lista")
    return render(request, "empleados/grado_form.html", {"form": form, "titulo": "Crear grado"})


@login_required
@user_passes_test(_can_access_backoffice)
def editar_grado(request, pk):
    grado = get_object_or_404(Grado, pk=pk)
    form = GradoForm(request.POST or None, instance=grado)
    if request.method == "POST" and form.is_valid():
        grado = form.save(commit=False)
        grado.carrera = grado.carrera
        grado.save()
        messages.success(request, "Grado actualizado.")
        return redirect("empleados:grado_lista")
    return render(request, "empleados/grado_form.html", {"form": form, "titulo": "Editar grado"})


@login_required
@user_passes_test(_can_access_backoffice)
def matricula_view(request):
    establecimiento_id = request.GET.get("establecimiento") or request.POST.get("establecimiento")
    carrera_id = request.GET.get("carrera") or request.POST.get("carrera")

    form = MatriculaForm(request.POST or None, establecimiento_id=establecimiento_id, carrera_id=carrera_id)
    if request.method == "POST" and form.is_valid():
        matricula = form.save()
        messages.success(request, "Matrícula registrada.")
        return redirect("empleados:matricula")

    matriculas = Matricula.objects.select_related("alumno", "grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento")
    matriculas = filtrar_por_establecimiento_usuario(matriculas, request.user, "grado__carrera__ciclo_escolar__establecimiento_id")
    grado_id = request.GET.get("grado")
    ciclo = request.GET.get("ciclo")
    ciclo_escolar_id = request.GET.get("ciclo_escolar")
    estado = request.GET.get("estado")
    if establecimiento_id:
        matriculas = matriculas.filter(grado__carrera__ciclo_escolar__establecimiento_id=establecimiento_id)
    if carrera_id:
        matriculas = matriculas.filter(grado__carrera_id=carrera_id)
    if grado_id:
        matriculas = matriculas.filter(grado_id=grado_id)
    if ciclo_escolar_id:
        matriculas = matriculas.filter(ciclo_escolar_id=ciclo_escolar_id)
    elif ciclo:
        matriculas = matriculas.filter(ciclo=ciclo)
    if estado:
        matriculas = matriculas.filter(estado=estado)

    establecimientos = Establecimiento.objects.filter(activo=True)
    establecimientos = filtrar_por_establecimiento_usuario(establecimientos, request.user, "id")
    carreras = Carrera.objects.filter(ciclo_escolar__establecimiento_id=establecimiento_id, activo=True) if establecimiento_id else Carrera.objects.none()
    grados = Grado.objects.filter(carrera_id=carrera_id, activo=True) if carrera_id else Grado.objects.none()

    return render(
        request,
        "empleados/matricula.html",
        {
            "form": form,
            "matriculas": matriculas,
            "establecimientos": establecimientos,
            "carreras": carreras,
            "grados": grados,
        },
    )


@login_required
@user_passes_test(_can_access_backoffice)
@require_GET
def matricula_masiva_buscar_alumnos(request):
    if _is_docente(request.user):
        return JsonResponse({"results": []}, status=403)

    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    alumnos = Empleado.objects.select_related("grado", "establecimiento")
    alumnos = filtrar_por_establecimiento_usuario(alumnos, request.user, "establecimiento_id")
    alumnos = alumnos.filter(Q(codigo_personal__icontains=q) | Q(nombres__icontains=q) | Q(apellidos__icontains=q)).order_by("codigo_personal", "apellidos", "nombres")[:15]

    results = [
        {
            "id": alumno.id,
            "codigo_personal": alumno.codigo_personal or "-",
            "nombre": f"{alumno.nombres} {alumno.apellidos}".strip(),
            "grado_actual": str(alumno.grado) if alumno.grado else "-",
            "establecimiento": str(alumno.establecimiento) if alumno.establecimiento else "-",
        }
        for alumno in alumnos
    ]
    return JsonResponse({"results": results})


@login_required
@user_passes_test(_can_access_backoffice)
def matricula_masiva(request):
    if _is_docente(request.user):
        return redirect("empleados:dahsboard")

    form = MatriculaMasivaForm(request.POST or None, user=request.user)
    establecimiento_usuario = obtener_establecimiento_usuario(request.user)

    if request.method == "POST" and form.is_valid():
        raw_ids = request.POST.get("alumnos_ids", "")
        alumno_ids = [int(v) for v in raw_ids.split(",") if v.strip().isdigit()]
        alumno_ids = list(dict.fromkeys(alumno_ids))

        if not alumno_ids:
            messages.warning(request, "Debe agregar al menos un alumno a la lista.")
            return render(request, "empleados/matricula_masiva.html", {"form": form})

        grado = form.cleaned_data["grado"]
        ciclo_escolar = form.cleaned_data["ciclo_escolar"]
        estado = form.cleaned_data["estado"]

        grado_establecimiento = grado.carrera.ciclo_escolar.establecimiento if grado.carrera else None
        if not grado_establecimiento or ciclo_escolar.establecimiento_id != grado_establecimiento.id:
            messages.error(request, "El ciclo escolar debe pertenecer al mismo establecimiento del grado seleccionado.")
            return render(request, "empleados/matricula_masiva.html", {"form": form})

        if establecimiento_usuario and grado_establecimiento.id != establecimiento_usuario.id:
            messages.error(request, "No tiene permisos para matricular en un establecimiento diferente al asignado.")
            return render(request, "empleados/matricula_masiva.html", {"form": form})

        alumnos_qs = Empleado.objects.filter(id__in=alumno_ids)
        alumnos_qs = filtrar_por_establecimiento_usuario(alumnos_qs, request.user, "establecimiento_id")
        alumnos_map = {a.id: a for a in alumnos_qs}

        inscritos = 0
        omitidos = 0
        errores = 0

        for alumno_id in alumno_ids:
            alumno = alumnos_map.get(alumno_id)
            if not alumno:
                errores += 1
                continue

            if Matricula.objects.filter(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar).exists():
                omitidos += 1
                continue

            matricula = Matricula(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar, estado=estado)
            try:
                matricula.full_clean()
                matricula.save()
                inscritos += 1
            except ValidationError:
                errores += 1

        if inscritos:
            messages.success(request, f"{inscritos} alumnos matriculados correctamente")
        if omitidos:
            messages.warning(request, f"{omitidos} alumnos ya tenían matrícula y fueron omitidos")
        if errores:
            messages.error(request, f"{errores} alumnos no pudieron matricularse por validación o permisos")

        return redirect("empleados:matricula_masiva")

    return render(request, "empleados/matricula_masiva.html", {"form": form})


def _load_excel_records(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise ValidationError("El archivo debe ser Excel (.xlsx o .xlsm).")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("No está disponible la librería para leer Excel en este entorno.") from exc

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
    data_rows = [list(r) for r in rows[1:] if any(c not in (None, "") for c in r)]
    return headers, data_rows


def _parse_date_cell(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    raise ValidationError("Fecha inválida. Use YYYY-MM-DD o DD/MM/YYYY.")


def _bool_cell(value, default=True):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    raw = str(value).strip().lower()
    return raw in {"1", "true", "si", "sí", "activo", "yes"}


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_home(request):
    return render(request, "empleados/carga_masiva/index.html")


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_plantilla(request, tipo):
    try:
        from openpyxl import Workbook
    except ImportError:
        messages.error(request, "No fue posible generar la plantilla en este entorno.")
        return redirect("empleados:carga_masiva_home")

    columnas_por_tipo = {
        "alumnos": ["codigo_personal", "nombres", "apellidos", "cui", "fecha_nacimiento", "telefono", "grado_id", "establecimiento_id", "activo"],
        "docentes": ["username", "first_name", "last_name", "email", "password", "activo"],
        "cursos": ["grado_id", "nombre", "descripcion", "activo"],
        "asignaciones": ["curso_id", "docente_username", "activo"],
    }
    columnas = columnas_por_tipo.get(tipo)
    if not columnas:
        messages.error(request, "Tipo de plantilla no válido.")
        return redirect("empleados:carga_masiva_home")

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.append(columnas)
    ws.append([""] * len(columnas))
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="plantilla_{tipo}.xlsx"'
    wb.save(response)
    return response


def _import_alumnos_excel(request, headers, rows, confirmar=False):
    requeridas = {"nombres", "apellidos"}
    faltantes = [c for c in requeridas if c not in headers]
    if faltantes:
        raise ValidationError(f"Columnas requeridas faltantes: {', '.join(faltantes)}")

    idx = {h: i for i, h in enumerate(headers)}
    grados_qs = filtrar_por_establecimiento_usuario(Grado.objects.select_related("carrera__ciclo_escolar__establecimiento"), request.user, "carrera__ciclo_escolar__establecimiento_id")
    grados_map = {g.id: g for g in grados_qs}
    establecimientos_qs = filtrar_por_establecimiento_usuario(Establecimiento.objects.all(), request.user, "id")
    establecimientos_map = {e.id: e for e in establecimientos_qs}

    operaciones = []
    errores = []
    for num, row in enumerate(rows, start=2):
        try:
            nombres = str(row[idx["nombres"]]).strip() if row[idx["nombres"]] else ""
            apellidos = str(row[idx["apellidos"]]).strip() if row[idx["apellidos"]] else ""
            if not nombres or not apellidos:
                raise ValidationError("Nombres y apellidos son obligatorios.")
            codigo = str(row[idx["codigo_personal"]]).strip() if "codigo_personal" in idx and row[idx["codigo_personal"]] else None
            cui = str(row[idx["cui"]]).strip() if "cui" in idx and row[idx["cui"]] else None
            fecha_nacimiento = _parse_date_cell(row[idx["fecha_nacimiento"]]) if "fecha_nacimiento" in idx else None
            tel = str(row[idx["telefono"]]).strip() if "telefono" in idx and row[idx["telefono"]] else None
            grado = None
            if "grado_id" in idx and row[idx["grado_id"]] not in (None, ""):
                grado_id = int(row[idx["grado_id"]])
                grado = grados_map.get(grado_id)
                if not grado:
                    raise ValidationError("grado_id no válido o sin permiso.")
            establecimiento = None
            if "establecimiento_id" in idx and row[idx["establecimiento_id"]] not in (None, ""):
                est_id = int(row[idx["establecimiento_id"]])
                establecimiento = establecimientos_map.get(est_id)
                if not establecimiento:
                    raise ValidationError("establecimiento_id no válido o sin permiso.")
            activo = _bool_cell(row[idx["activo"]], default=True) if "activo" in idx else True

            existing = None
            if codigo:
                existing = Empleado.objects.filter(codigo_personal=codigo).first()
            if not existing and cui:
                existing = Empleado.objects.filter(cui=cui).first()
            operaciones.append({
                "existing": existing,
                "payload": {
                    "codigo_personal": codigo,
                    "nombres": nombres,
                    "apellidos": apellidos,
                    "cui": cui,
                    "fecha_nacimiento": fecha_nacimiento,
                    "tel": tel,
                    "grado": grado,
                    "establecimiento": establecimiento,
                    "activo": activo,
                    "user": request.user,
                },
            })
        except Exception as exc:  # validación por fila
            errores.append(f"Fila {num}: {exc}")

    creados = actualizados = omitidos = 0
    if confirmar and not errores:
        with transaction.atomic():
            for op in operaciones:
                if op["existing"]:
                    empleado = op["existing"]
                    for key, value in op["payload"].items():
                        setattr(empleado, key, value)
                    empleado.save()
                    actualizados += 1
                else:
                    Empleado.objects.create(**op["payload"])
                    creados += 1
    else:
        for op in operaciones:
            if op["existing"]:
                actualizados += 1
            else:
                creados += 1
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores, "preview": operaciones[:20]}


def _import_docentes_excel(headers, rows, confirmar=False):
    if "username" not in headers:
        raise ValidationError("La columna username es obligatoria.")
    idx = {h: i for i, h in enumerate(headers)}
    docentes_group, _ = Group.objects.get_or_create(name="Docente")

    operaciones, errores = [], []
    for num, row in enumerate(rows, start=2):
        try:
            username = str(row[idx["username"]]).strip() if row[idx["username"]] else ""
            if not username:
                raise ValidationError("username obligatorio.")
            first_name = str(row[idx["first_name"]]).strip() if "first_name" in idx and row[idx["first_name"]] else ""
            last_name = str(row[idx["last_name"]]).strip() if "last_name" in idx and row[idx["last_name"]] else ""
            email = str(row[idx["email"]]).strip() if "email" in idx and row[idx["email"]] else ""
            password = str(row[idx["password"]]).strip() if "password" in idx and row[idx["password"]] else None
            activo = _bool_cell(row[idx["activo"]], default=True) if "activo" in idx else True
            user = User.objects.filter(username=username).first()
            operaciones.append({"existing": user, "payload": {"username": username, "first_name": first_name, "last_name": last_name, "email": email, "password": password, "is_active": activo}})
        except Exception as exc:
            errores.append(f"Fila {num}: {exc}")

    creados = actualizados = omitidos = 0
    if confirmar and not errores:
        with transaction.atomic():
            for op in operaciones:
                payload = op["payload"]
                if op["existing"]:
                    user = op["existing"]
                    user.first_name = payload["first_name"]
                    user.last_name = payload["last_name"]
                    user.email = payload["email"]
                    user.is_active = payload["is_active"]
                    if payload["password"]:
                        user.set_password(payload["password"])
                    user.save()
                    actualizados += 1
                else:
                    user = User.objects.create_user(
                        username=payload["username"],
                        password=payload["password"] or "Docente123*",
                        first_name=payload["first_name"],
                        last_name=payload["last_name"],
                        email=payload["email"],
                        is_active=payload["is_active"],
                    )
                    creados += 1
                user.groups.add(docentes_group)
    else:
        for op in operaciones:
            if op["existing"]:
                actualizados += 1
            else:
                creados += 1
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores, "preview": operaciones[:20]}


def _import_cursos_excel(request, headers, rows, confirmar=False):
    if "grado_id" not in headers or "nombre" not in headers:
        raise ValidationError("Las columnas grado_id y nombre son obligatorias.")
    idx = {h: i for i, h in enumerate(headers)}
    grados_qs = filtrar_por_establecimiento_usuario(Grado.objects.all(), request.user, "carrera__ciclo_escolar__establecimiento_id")
    grados_map = {g.id: g for g in grados_qs}

    operaciones, errores = [], []
    for num, row in enumerate(rows, start=2):
        try:
            grado_id = int(row[idx["grado_id"]])
            grado = grados_map.get(grado_id)
            if not grado:
                raise ValidationError("grado_id no válido o sin permiso.")
            nombre = str(row[idx["nombre"]]).strip() if row[idx["nombre"]] else ""
            if not nombre:
                raise ValidationError("nombre obligatorio.")
            descripcion = str(row[idx["descripcion"]]).strip() if "descripcion" in idx and row[idx["descripcion"]] else ""
            activo = _bool_cell(row[idx["activo"]], default=True) if "activo" in idx else True
            existing = Curso.objects.filter(grado=grado, nombre=nombre).first()
            operaciones.append({"existing": existing, "payload": {"grado": grado, "nombre": nombre, "descripcion": descripcion, "activo": activo}})
        except Exception as exc:
            errores.append(f"Fila {num}: {exc}")

    creados = actualizados = omitidos = 0
    if confirmar and not errores:
        with transaction.atomic():
            for op in operaciones:
                if op["existing"]:
                    curso = op["existing"]
                    curso.descripcion = op["payload"]["descripcion"]
                    curso.activo = op["payload"]["activo"]
                    curso.save()
                    actualizados += 1
                else:
                    Curso.objects.create(**op["payload"])
                    creados += 1
    else:
        for op in operaciones:
            if op["existing"]:
                actualizados += 1
            else:
                creados += 1
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores, "preview": operaciones[:20]}


def _import_asignaciones_excel(request, headers, rows, confirmar=False):
    if "curso_id" not in headers or "docente_username" not in headers:
        raise ValidationError("Las columnas curso_id y docente_username son obligatorias.")
    idx = {h: i for i, h in enumerate(headers)}
    cursos_qs = filtrar_por_establecimiento_usuario(Curso.objects.all(), request.user, "grado__carrera__ciclo_escolar__establecimiento_id")
    cursos_map = {c.id: c for c in cursos_qs}
    docentes = User.objects.filter(groups__name="Docente").distinct()
    docentes_map = {u.username: u for u in docentes}

    operaciones, errores = [], []
    for num, row in enumerate(rows, start=2):
        try:
            curso = cursos_map.get(int(row[idx["curso_id"]]))
            if not curso:
                raise ValidationError("curso_id no válido o sin permiso.")
            username = str(row[idx["docente_username"]]).strip() if row[idx["docente_username"]] else ""
            docente = docentes_map.get(username)
            if not docente:
                raise ValidationError("docente_username no existe o no pertenece al grupo Docente.")
            activo = _bool_cell(row[idx["activo"]], default=True) if "activo" in idx else True
            existing = CursoDocente.objects.filter(curso=curso, docente=docente).first()
            operaciones.append({"existing": existing, "payload": {"curso": curso, "docente": docente, "activo": activo}})
        except Exception as exc:
            errores.append(f"Fila {num}: {exc}")

    creados = actualizados = omitidos = 0
    if confirmar and not errores:
        with transaction.atomic():
            for op in operaciones:
                if op["existing"]:
                    op["existing"].activo = op["payload"]["activo"]
                    op["existing"].save(update_fields=["activo"])
                    actualizados += 1
                else:
                    CursoDocente.objects.create(**op["payload"])
                    creados += 1
    else:
        for op in operaciones:
            if op["existing"]:
                actualizados += 1
            else:
                creados += 1
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores, "preview": operaciones[:20]}


def _carga_masiva_import_view(request, tipo, titulo, descripcion):
    form = CargaMasivaExcelForm(request.POST or None, request.FILES or None)
    resultado = None

    if request.method == "POST" and form.is_valid():
        try:
            headers, rows = _load_excel_records(form.cleaned_data["archivo"])
            confirmar = bool(form.cleaned_data.get("confirmar"))
            if tipo == "alumnos":
                resultado = _import_alumnos_excel(request, headers, rows, confirmar=confirmar)
            elif tipo == "docentes":
                resultado = _import_docentes_excel(headers, rows, confirmar=confirmar)
            elif tipo == "cursos":
                resultado = _import_cursos_excel(request, headers, rows, confirmar=confirmar)
            elif tipo == "asignaciones":
                resultado = _import_asignaciones_excel(request, headers, rows, confirmar=confirmar)
            else:
                raise ValidationError("Tipo de importación no soportado.")

            if confirmar and not resultado["errores"]:
                messages.success(
                    request,
                    f"Importación completada. Creados: {resultado['creados']}, actualizados: {resultado['actualizados']}, omitidos: {resultado['omitidos']}.",
                )
            elif resultado["errores"]:
                messages.warning(request, f"Se detectaron {len(resultado['errores'])} errores. Corrija y vuelva a intentar.")
            else:
                messages.info(request, "Previsualización generada. Marque confirmar para aplicar cambios.")
        except ValidationError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            messages.error(request, f"No fue posible procesar el archivo: {exc}")

    return render(
        request,
        "empleados/carga_masiva/importar.html",
        {
            "form": form,
            "titulo": titulo,
            "descripcion": descripcion,
            "tipo": tipo,
            "resultado": resultado,
        },
    )


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_import_alumnos(request):
    return _carga_masiva_import_view(request, "alumnos", "Importar alumnos", "Carga de alumnos desde Excel.")


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_import_docentes(request):
    return _carga_masiva_import_view(request, "docentes", "Importar docentes", "Carga de usuarios docentes desde Excel.")


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_import_cursos(request):
    return _carga_masiva_import_view(request, "cursos", "Importar cursos", "Carga de cursos vinculados a grado desde Excel.")


@login_required
@user_passes_test(_can_access_backoffice)
def carga_masiva_import_asignaciones(request):
    return _carga_masiva_import_view(request, "asignaciones", "Importar asignaciones", "Carga de asignaciones docente-curso desde Excel.")


@login_required
@user_passes_test(_can_access_backoffice)
def editor_gafete(request, establecimiento_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    denied = _deny_if_not_allowed_establecimiento(request, establecimiento_id)
    if denied:
        return denied
    establecimiento = get_object_or_404(Establecimiento, pk=establecimiento_id)
    matricula_demo = (
        Matricula.objects.select_related("alumno", "grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento")
        .filter(grado__carrera__ciclo_escolar__establecimiento=establecimiento, estado="activo")
        .order_by("-created_at")
        .first()
    )
    alumno = matricula_demo.alumno if matricula_demo else Empleado.objects.first()
    grado_demo = matricula_demo.grado if matricula_demo else None
    orientation = orientation_for_establecimiento(establecimiento)
    layout = normalizar_layout_gafete(establecimiento.get_layout(), orientation=orientation)
    canvas_width, canvas_height = canvas_for_orientation(orientation)
    layout["canvas"] = {"width": canvas_width, "height": canvas_height, "orientation": orientation}
    available_fields = [
        {"key": "photo", "label": "Foto"},
        {"key": "nombres", "label": "Nombres"},
        {"key": "apellidos", "label": "Apellidos"},
        {"key": "codigo_alumno", "label": "Código alumno"},
        {"key": "grado", "label": "Grado"},
        {"key": "grado_descripcion", "label": "Descripción grado"},
        {"key": "cui", "label": "CUI"},
        {"key": "telefono", "label": "Teléfono emergencia"},
        {"key": "establecimiento", "label": "Establecimiento"},
        {"key": "sitio_web", "label": "Sitio web"},
        {"key": "texto_libre_1", "label": "Texto libre 1"},
        {"key": "texto_libre_2", "label": "Texto libre 2"},
        {"key": "texto_libre_3", "label": "Texto libre 3"},
    ]
    configuracion = ConfiguracionGeneral.objects.first()
    return render(
        request,
        "aulapro/establecimiento_gafete_editor.html",
        {
            "establecimiento": establecimiento,
            "alumno": alumno,
            "grado_demo": grado_demo,
            "layout": layout,
            "layout_json": json.dumps(layout),
            "default_layout_json": json.dumps(DEFAULT_GAFETE_LAYOUT),
            "available_fields": available_fields,
            "enabled_fields_front": obtener_layout_cara(layout, "front").get("enabled_fields", []),
            "enabled_fields_back": obtener_layout_cara(layout, "back").get("enabled_fields", []),
            "configuracion": configuracion,
            "is_editor": True,
            "canvas_width": canvas_width,
            "canvas_height": canvas_height,
            "gafete_w": canvas_width,
            "gafete_h": canvas_height,
            "orientacion": orientation,
        },
    )




@login_required
@user_passes_test(_can_access_backoffice)
@require_POST
def subir_imagen_gafete(request, establecimiento_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    denied = _deny_if_not_allowed_establecimiento(request, establecimiento_id)
    if denied:
        return denied
    _ = get_object_or_404(Establecimiento, pk=establecimiento_id)
    image = request.FILES.get("image")
    if not image:
        return JsonResponse({"ok": False, "error": "Debe seleccionar una imagen."}, status=400)
    if image.size > 4 * 1024 * 1024:
        return JsonResponse({"ok": False, "error": "La imagen supera 4MB."}, status=400)
    if image.content_type not in {"image/jpeg", "image/png", "image/webp", "image/svg+xml"}:
        return JsonResponse({"ok": False, "error": "Formato no permitido."}, status=400)
    try:
        stored_path = default_storage.save(f"gafetes/overlays/{uuid4().hex}_{image.name}", image)
    except Exception:
        return JsonResponse({"ok": False, "error": "No se pudo almacenar la imagen."}, status=500)
    return JsonResponse({"ok": True, "url": default_storage.url(stored_path)})


@login_required
@user_passes_test(_can_access_backoffice)
@require_POST
def guardar_diseno_gafete(request, establecimiento_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    denied = _deny_if_not_allowed_establecimiento(request, establecimiento_id)
    if denied:
        return denied
    establecimiento = get_object_or_404(Establecimiento, pk=establecimiento_id)
    try:
        payload = json.loads(request.body.decode("utf-8"))
        orientation = orientation_for_establecimiento(establecimiento)
        layout = _validate_layout_payload(payload, forced_orientation=orientation)
    except (ValueError, json.JSONDecodeError, TypeError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)
    orientation = orientation_for_establecimiento(establecimiento)
    canvas_width, canvas_height = canvas_for_orientation(orientation)
    layout = serializar_layout_frente_reverso(layout, orientation=orientation)
    layout["canvas"] = {"width": canvas_width, "height": canvas_height, "orientation": orientation}
    establecimiento.gafete_ancho = canvas_width
    establecimiento.gafete_alto = canvas_height
    establecimiento.gafete_layout_json = layout
    establecimiento.save(update_fields=["gafete_layout_json", "gafete_ancho", "gafete_alto"])
    if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.content_type == "application/json":
        return JsonResponse({"ok": True})
    messages.success(request, "Diseño guardado correctamente.")
    return redirect("empleados:editor_gafete", establecimiento_id=establecimiento.id)



def _safe_text(value, fallback="-"):
    text = str(value or "").strip()
    return text if text else fallback


def _sanitize_filename_token(value):
    text = unicodedata.normalize("NFKD", _safe_text(value, fallback=""))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text or "NA"


def _build_gafete_filename(alumno, lado="frente"):
    apellidos = _sanitize_filename_token(getattr(alumno, "apellidos", ""))
    nombres = _sanitize_filename_token(getattr(alumno, "nombres", ""))
    codigo = _sanitize_filename_token(getattr(alumno, "codigo_personal", ""))
    lado_token = _sanitize_filename_token(lado)
    return f"GAFETE_{lado_token}_{apellidos}_{nombres}_{codigo}.jpg"


def _parse_color(value, default="#111111"):
    try:
        return ImageColor.getrgb(str(value or default))
    except ValueError:
        return ImageColor.getrgb(default)


def _load_font(font_size=24, bold=False):
    font_candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    for font_path in font_candidates:
        try:
            return ImageFont.truetype(font_path, max(10, int(font_size or 24)))
        except OSError:
            continue
    return ImageFont.load_default()


def _field_text_for_key(key, matricula, establecimiento):
    alumno = matricula.alumno
    grado = matricula.grado
    mapping = {
        "nombres": _safe_text(getattr(alumno, "nombres", ""), "Nombre"),
        "apellidos": _safe_text(getattr(alumno, "apellidos", ""), "Apellidos"),
        "codigo_alumno": _safe_text(getattr(alumno, "codigo_personal", ""), "-"),
        "grado": _safe_text(getattr(grado, "nombre", ""), "Grado"),
        "grado_descripcion": _safe_text(getattr(grado, "descripcion", ""), ""),
        "sitio_web": _safe_text(getattr(establecimiento, "sitio_web", ""), ""),
        "telefono": _safe_text(getattr(alumno, "tel", ""), "-"),
        "cui": _safe_text(getattr(alumno, "cui", ""), "—"),
        "establecimiento": _safe_text(getattr(establecimiento, "nombre", ""), ""),
    }
    return mapping.get(key, "")




def _resolve_media_source(path_or_url):
    source = str(path_or_url or '').strip()
    if not source:
        return None
    if source.startswith(('http://', 'https://')):
        return source
    if source.startswith('/media/'):
        return default_storage.path(source.replace('/media/', '', 1))
    if source.startswith('media/'):
        return default_storage.path(source.replace('media/', '', 1))
    return source

def _apply_cover_image(src_image, target_w, target_h):
    return ImageOps.fit(src_image, (target_w, target_h), method=Image.Resampling.LANCZOS)


def _apply_contain_image(src_image, target_w, target_h):
    contained = ImageOps.contain(src_image, (target_w, target_h), method=Image.Resampling.LANCZOS)
    layer = Image.new("RGBA", (target_w, target_h), (255, 255, 255, 0))
    x = (target_w - contained.width) // 2
    y = (target_h - contained.height) // 2
    layer.paste(contained, (x, y))
    return layer


def _open_normalized_image(file_obj):
    image = Image.open(file_obj)
    image = ImageOps.exif_transpose(image)
    return image


def _draw_wrapped_text(draw, text, x, y, max_w, max_h, font, fill, align="left"):
    words = str(text or "").split()
    if not words:
        return
    lines = []
    line = words[0]
    for word in words[1:]:
        candidate = f"{line} {word}".strip()
        bbox = draw.textbbox((0, 0), candidate, font=font)
        if bbox[2] - bbox[0] <= max_w:
            line = candidate
        else:
            lines.append(line)
            line = word
    lines.append(line)

    line_h = draw.textbbox((0, 0), "Ag", font=font)[3] + 2
    max_lines = max(1, int(max_h // line_h))
    lines = lines[:max_lines]
    for idx, ln in enumerate(lines):
        tw = draw.textbbox((0, 0), ln, font=font)[2]
        tx = x
        if align == "center":
            tx = x + max((max_w - tw) // 2, 0)
        elif align == "right":
            tx = x + max(max_w - tw, 0)
        draw.text((tx, y + idx * line_h), ln, fill=fill, font=font)


def renderizar_elementos_gafete(canvas, matricula, establecimiento, face_layout, face="front"):
    items = face_layout.get("items", {}) if isinstance(face_layout, dict) else {}

    photo_cfg = items.get("photo", {}) if isinstance(items.get("photo", {}), dict) else {}
    if is_item_visible_in_face(face_layout, face, "photo") and getattr(matricula.alumno, "imagen", None):
        x = int(photo_cfg.get("x", 20))
        y = int(photo_cfg.get("y", 40))
        w = max(20, int(photo_cfg.get("w", 250)))
        h = max(20, int(photo_cfg.get("h", 350)) )
        border_width = max(0, int(photo_cfg.get("border_width", 4))) if photo_cfg.get("border", True) else 0
        border_color = _parse_color(photo_cfg.get("border_color", "#ffffff"), default="#ffffff")
        shape = str(photo_cfg.get("shape") or "rounded").lower()
        radius = max(0, int(photo_cfg.get("radius", 20)))
        try:
            with matricula.alumno.imagen.open("rb") as photo_file:
                photo = _open_normalized_image(photo_file).convert("RGB")
                photo = _apply_cover_image(photo, w, h)
                alpha_mask = Image.new("L", (w, h), 0)
                alpha_draw = ImageDraw.Draw(alpha_mask)
                if shape == "circle":
                    alpha_draw.ellipse((0, 0, w, h), fill=255)
                else:
                    alpha_draw.rounded_rectangle((0, 0, w, h), radius=radius, fill=255)
                if border_width > 0:
                    border_img = Image.new("RGBA", (w + border_width * 2, h + border_width * 2), (0, 0, 0, 0))
                    border_mask = Image.new("L", border_img.size, 0)
                    border_draw = ImageDraw.Draw(border_mask)
                    if shape == "circle":
                        border_draw.ellipse((0, 0, border_img.size[0], border_img.size[1]), fill=255)
                    else:
                        border_draw.rounded_rectangle((0, 0, border_img.size[0], border_img.size[1]), radius=radius + border_width, fill=255)
                    border_fill = Image.new("RGBA", border_img.size, (*border_color, 255))
                    border_img.paste(border_fill, (0, 0), border_mask)
                    canvas.paste(border_img, (x - border_width, y - border_width), border_img)
                photo_rgba = photo.convert("RGBA")
                photo_rgba.putalpha(alpha_mask)
                canvas.paste(photo_rgba, (x, y), photo_rgba)
        except Exception:
            pass

    image_keys = [key for key in items.keys() if str(key).startswith("image")]
    for image_key in image_keys:
        image_cfg = items.get(image_key, {}) if isinstance(items.get(image_key, {}), dict) else {}
        if not is_item_visible_in_face(face_layout, face, image_key) or not image_cfg.get("src"):
            continue
        try:
            from urllib.request import urlopen
            img_src = str(image_cfg.get("src"))
            resolved_source = _resolve_media_source(img_src)
            if str(resolved_source).startswith(("http://", "https://")):
                raw = urlopen(resolved_source).read()
                overlay = Image.open(BytesIO(raw)).convert("RGBA")
            else:
                with open(resolved_source, "rb") as image_file:
                    overlay = Image.open(image_file).convert("RGBA")
            w = max(20, int(image_cfg.get("w", 220)))
            h = max(20, int(image_cfg.get("h", 220)))
            prepared = _apply_contain_image(overlay, w, h) if image_cfg.get("object_fit") == "contain" else _apply_cover_image(overlay, w, h).convert("RGBA")
            canvas.paste(prepared, (int(image_cfg.get("x", 30)), int(image_cfg.get("y", 30))), prepared)
        except Exception:
            pass

    draw = ImageDraw.Draw(canvas)
    for key, cfg in items.items():
        if str(key).startswith("image") or key == "photo" or not isinstance(cfg, dict) or not is_item_visible_in_face(face_layout, face, key):
            continue
        text = cfg.get("text") if key.startswith("texto_libre_") else _field_text_for_key(key, matricula, establecimiento)
        if not text:
            continue
        x = int(cfg.get("x", 0))
        y = int(cfg.get("y", 0))
        font_size = int(cfg.get("font_size", 24))
        weight = str(cfg.get("font_weight", "400"))
        color = _parse_color(cfg.get("color", "#111111"), default="#111111")
        align = str(cfg.get("align", "left")).lower()
        font = _load_font(font_size=font_size, bold=(weight == "700"))
        if "w" in cfg and "h" in cfg:
            max_w = max(20, int(cfg.get("w", 280)))
            max_h = max(20, int(cfg.get("h", 70)))
            _draw_wrapped_text(draw, text, x, y, max_w, max_h, font, color, align=align)
        else:
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_w = text_bbox[2] - text_bbox[0]
            tx = x - text_w // 2 if align == "center" else x - text_w if align == "right" else x
            draw.text((tx, y), text, fill=color, font=font)


def _render_face_gafete(matricula, establecimiento, layout, face, canvas_width, canvas_height):
    canvas = Image.new("RGB", (canvas_width, canvas_height), "white")
    face_layout = obtener_layout_cara(layout, face)
    bg_url = face_layout.get("background_image") if isinstance(face_layout, dict) else ""
    if face == "back" and establecimiento and getattr(establecimiento, "background_gafete_posterior", None):
        bg_url = establecimiento.background_gafete_posterior.url

    if bg_url:
        try:
            from urllib.request import urlopen
            resolved_bg = _resolve_media_source(bg_url)
            raw = urlopen(resolved_bg).read() if str(resolved_bg).startswith(("http://", "https://")) else open(resolved_bg, "rb").read()
            background = Image.open(BytesIO(raw)).convert("RGB")
            canvas.paste(_apply_cover_image(background, canvas_width, canvas_height), (0, 0))
        except Exception:
            pass
    elif establecimiento and establecimiento.background_gafete and face == "front":
        try:
            with establecimiento.background_gafete.open("rb") as bg_file:
                background = Image.open(bg_file).convert("RGB")
                canvas.paste(_apply_cover_image(background, canvas_width, canvas_height), (0, 0))
        except Exception:
            pass

    renderizar_elementos_gafete(canvas, matricula, establecimiento, face_layout, face=face)

    if face == "front":
        config = ConfiguracionGeneral.objects.first()
        if config and config.logotipo:
            try:
                with config.logotipo.open("rb") as logo_file:
                    logo = Image.open(logo_file).convert("RGBA")
                    logo.thumbnail((170, 170), Image.Resampling.LANCZOS)
                    canvas.paste(logo, (12, 40), logo)
            except Exception:
                pass
    return canvas


def generar_descarga_gafete_alumno(matricula, establecimiento, layout, canvas_width, canvas_height, lado="frente"):
    front = _render_face_gafete(matricula, establecimiento, layout, "front", canvas_width, canvas_height)
    back = _render_face_gafete(matricula, establecimiento, layout, "back", canvas_width, canvas_height)

    if lado == "frente":
        output = front
    else:
        output = back

    buffer = BytesIO()
    output.save(buffer, format="JPEG", quality=95, optimize=True)
    return buffer.getvalue()


def _build_gafete_download_context(matricula, lado):
    establecimiento = matricula.grado.carrera.ciclo_escolar.establecimiento if matricula.grado and matricula.grado.carrera else None
    if not establecimiento:
        return None
    orientation = orientation_for_establecimiento(establecimiento)
    layout = normalizar_layout_gafete(establecimiento.get_layout() if establecimiento else DEFAULT_GAFETE_LAYOUT, orientation=orientation)
    canvas_width, canvas_height = canvas_for_orientation(orientation)
    face = "back" if lado == "reverso" else "front"
    return {
        "matricula": matricula,
        "alumno": matricula.alumno,
        "grado": matricula.grado,
        "establecimiento": establecimiento,
        "configuracion": ConfiguracionGeneral.objects.first(),
        "layout": layout,
        "face": face,
        "face_layout": obtener_layout_cara(layout, face),
        "canvas_width": canvas_width,
        "canvas_height": canvas_height,
        "gafete_w": canvas_width,
        "gafete_h": canvas_height,
        "filename": _build_gafete_filename(matricula.alumno, lado=lado),
        "lado": lado,
    }


@login_required
@user_passes_test(_can_access_backoffice)
def gafete_jpg(request, matricula_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    matricula = get_object_or_404(
        Matricula.objects.select_related("alumno", "grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento"),
        pk=matricula_id,
    )
    establecimiento = matricula.grado.carrera.ciclo_escolar.establecimiento if matricula.grado and matricula.grado.carrera else None
    if establecimiento:
        denied = _deny_if_not_allowed_establecimiento(request, establecimiento.id)
        if denied:
            return denied
    if not establecimiento:
        return HttpResponse("No se encontró establecimiento para la matrícula.", status=404)

    layout = normalizar_layout_gafete(establecimiento.get_layout() if establecimiento else DEFAULT_GAFETE_LAYOUT, orientation=orientation_for_establecimiento(establecimiento))
    orientation = orientation_for_establecimiento(establecimiento)
    canvas_width, canvas_height = canvas_for_orientation(orientation)
    lado = (request.GET.get("lado") or "frente").strip().lower()
    if lado not in {"frente", "reverso"}:
        lado = "frente"
    filename = _build_gafete_filename(matricula.alumno, lado=lado)
    return render(
        request,
        "aulapro/gafete_download.html",
        {
            "alumno": matricula.alumno,
            "grado": matricula.grado,
            "establecimiento": establecimiento,
            "layout": layout,
            "canvas_width": canvas_width,
            "canvas_height": canvas_height,
            "gafete_w": canvas_width,
            "gafete_h": canvas_height,
            "lado": lado,
            "download_filename": filename,
        },
    )


@login_required
@user_passes_test(_can_access_backoffice)
def descargar_gafete_jpg(request, matricula_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    return gafete_jpg(request, matricula_id)


@login_required
@user_passes_test(_can_access_backoffice)
def descargar_gafete_frente_jpg(request, matricula_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    matricula = get_object_or_404(
        Matricula.objects.select_related("alumno", "grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento"),
        pk=matricula_id,
    )
    establecimiento = matricula.grado.carrera.ciclo_escolar.establecimiento if matricula.grado and matricula.grado.carrera else None
    if establecimiento:
        denied = _deny_if_not_allowed_establecimiento(request, establecimiento.id)
        if denied:
            return denied
    ctx = _build_gafete_download_context(matricula, "frente")
    if not ctx:
        return HttpResponse("No se encontró establecimiento para la matrícula.", status=404)
    return render(request, "aulapro/gafete_download_face.html", ctx)


@login_required
@user_passes_test(_can_access_backoffice)
def descargar_gafete_reverso_jpg(request, matricula_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    matricula = get_object_or_404(
        Matricula.objects.select_related("alumno", "grado", "grado__carrera", "grado__carrera__ciclo_escolar__establecimiento"),
        pk=matricula_id,
    )
    establecimiento = matricula.grado.carrera.ciclo_escolar.establecimiento if matricula.grado and matricula.grado.carrera else None
    if establecimiento:
        denied = _deny_if_not_allowed_establecimiento(request, establecimiento.id)
        if denied:
            return denied
    ctx = _build_gafete_download_context(matricula, "reverso")
    if not ctx:
        return HttpResponse("No se encontró establecimiento para la matrícula.", status=404)
    return render(request, "aulapro/gafete_download_face.html", ctx)


@login_required
@user_passes_test(_can_access_backoffice)
def resetear_diseno_gafete(request, establecimiento_id):
    forbidden = _forbid_gafetes_for_gestor(request)
    if forbidden:
        return forbidden
    denied = _deny_if_not_allowed_establecimiento(request, establecimiento_id)
    if denied:
        return denied
    establecimiento = get_object_or_404(Establecimiento, pk=establecimiento_id)
    establecimiento.gafete_layout_json = {}
    establecimiento.gafete_ancho, establecimiento.gafete_alto = canvas_for_orientation("H")
    establecimiento.save(update_fields=["gafete_layout_json", "gafete_ancho", "gafete_alto"])
    messages.success(request, "Diseño restablecido al valor original.")
    return redirect("empleados:editor_gafete", establecimiento_id=establecimiento.id)
