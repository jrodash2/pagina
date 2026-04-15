from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_FONT = Font(size=14, bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_title(ws, row, text, max_col=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_table_header(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_table_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top")


def autosize_columns(ws, min_width=12, max_width=45):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def workbook_to_response(workbook, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response
