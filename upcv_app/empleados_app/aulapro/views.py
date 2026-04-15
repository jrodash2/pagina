import base64
import binascii
import io
import json
import os
import uuid
from django import forms
import logging
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Case, Count, F, IntegerField, Q, Sum, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from PIL import Image, ImageOps, UnidentifiedImageError

from openpyxl import Workbook
from openpyxl.styles import Font
from django.views.decorators.http import require_GET, require_POST

from empleados_app.forms import AsignarDocenteForm, CarreraForm, CicloEscolarForm, CursoForm, EstablecimientoForm, GradoForm
from empleados_app.gafete_utils import resolve_gafete_dimensions
from empleados_app.models import Asistencia, AsistenciaDetalle, Carrera, CicloEscolar, ConfiguracionGeneral, Curso, CursoDocente, Empleado, Establecimiento, Grado, Matricula, Perfil, PeriodoAcademico
from empleados_app.permissions import (
    es_admin_total,
    es_docente,
    es_gestor,
    filtrar_por_establecimiento_usuario,
    obtener_establecimiento_usuario,
    usuario_puede_ver_establecimiento,
)

from .excel import autosize_columns, style_table_header, style_table_row, style_title, workbook_to_response
from .forms import MatriculaFiltroForm

ALLOW_MULTI_GRADE_PER_CYCLE = False

logger = logging.getLogger(__name__)

MAX_ALUMNO_IMAGE_SIZE = 5 * 1024 * 1024
ALUMNO_IMAGE_MIME_EXT = {
    "image/jpeg": "jpg",
    "image/png": "png",
    "image/webp": "webp",
}



BASE_GAFETE_W = 1011
BASE_GAFETE_H = 639


def _canvas_for_orientation(orientation):
    return (BASE_GAFETE_W, BASE_GAFETE_H) if orientation == 'H' else (BASE_GAFETE_H, BASE_GAFETE_W)


def _resolve_gafete_dimensions(establecimiento, layout):
    orientation = str((layout or {}).get('canvas', {}).get('orientation') or ('V' if (establecimiento.gafete_alto or 0) > (establecimiento.gafete_ancho or 0) else 'H')).upper()
    if orientation not in ('H', 'V'):
        orientation = 'H'
    gafete_w, gafete_h = _canvas_for_orientation(orientation)
    return orientation, gafete_w, gafete_h


def _can_manage(user):
    return es_admin_total(user) or es_gestor(user)


def _is_docente(user):
    return es_docente(user)


def _can_view_attendance(user):
    return _is_docente(user) or _can_manage(user)


def _attendance_filter_for_user(user, prefix=""):
    if _is_docente(user):
        return {
            f"{prefix}docente": user,
            f"{prefix}curso__grado__carrera__ciclo_escolar__activo": True,
        }
    if es_gestor(user):
        establecimiento = obtener_establecimiento_usuario(user)
        if not establecimiento:
            return {f"{prefix}pk__in": []}
        return {f"{prefix}curso__grado__carrera__ciclo_escolar__establecimiento_id": establecimiento.id}
    return {}


def _ensure_establecimiento_access(request, est_id):
    if not usuario_puede_ver_establecimiento(request.user, est_id):
        messages.error(request, 'No tiene permisos para ese establecimiento.')
        return redirect('empleados:dahsboard')
    return None




def _gestores_qs_para_establecimiento(establecimiento):
    return (
        User.objects.filter(groups__name="Gestor", perfil__establecimiento_gestionado=establecimiento)
        .select_related("perfil")
        .order_by("first_name", "last_name", "username")
        .distinct()
    )


def _gestores_disponibles_qs():
    return User.objects.filter(groups__name="Gestor").order_by("first_name", "last_name", "username").distinct()


def _asignar_gestor_a_establecimiento(request, establecimiento):
    if not es_admin_total(request.user):
        messages.error(request, 'Solo un administrador puede gestionar asignaciones de gestores.')
        return

    action = (request.POST.get('action') or '').strip()
    gestor_id = request.POST.get('gestor_id')

    if not gestor_id:
        messages.warning(request, 'Debe seleccionar un gestor válido.')
        return

    gestor = get_object_or_404(_gestores_disponibles_qs(), pk=gestor_id)
    perfil, _ = Perfil.objects.get_or_create(user=gestor)

    if action == 'assign_gestor':
        anterior = perfil.establecimiento_gestionado
        perfil.establecimiento_gestionado = establecimiento
        perfil.save(update_fields=['establecimiento_gestionado'])
        if anterior and anterior.id != establecimiento.id:
            messages.success(request, f'Gestor reasignado desde "{anterior.nombre}" hacia "{establecimiento.nombre}".')
        else:
            messages.success(request, 'Gestor asignado correctamente al establecimiento.')
        return

    if action == 'unassign_gestor':
        if perfil.establecimiento_gestionado_id != establecimiento.id:
            messages.warning(request, 'El gestor seleccionado no está asignado a este establecimiento.')
            return
        perfil.establecimiento_gestionado = None
        perfil.save(update_fields=['establecimiento_gestionado'])
        messages.success(request, 'Gestor desasignado correctamente.')
        return

    messages.warning(request, 'Acción de gestor no reconocida.')

def _display_name_for_person(person):
    if not person:
        return ''

    get_full_name = getattr(person, 'get_full_name', None)
    if callable(get_full_name):
        full_name = (get_full_name() or '').strip()
        if full_name:
            return full_name

    nombres = getattr(person, 'nombres', '') or ''
    apellidos = getattr(person, 'apellidos', '') or ''
    nombre_empleado = f'{nombres} {apellidos}'.strip()
    if nombre_empleado:
        return nombre_empleado

    first_name = getattr(person, 'first_name', '') or ''
    last_name = getattr(person, 'last_name', '') or ''
    nombre_usuario = f'{first_name} {last_name}'.strip()
    if nombre_usuario:
        return nombre_usuario

    username = getattr(person, 'username', '') or ''
    if username:
        return username

    return str(person)




def _get_previous_cycle_for_establecimiento(ciclo_nuevo):
    return (
        CicloEscolar.objects.filter(establecimiento=ciclo_nuevo.establecimiento)
        .exclude(pk=ciclo_nuevo.pk)
        .order_by('-anio', '-id')
        .first()
    )


def _clone_academic_structure_from_previous_cycle(ciclo_nuevo):
    ciclo_anterior = _get_previous_cycle_for_establecimiento(ciclo_nuevo)
    if not ciclo_anterior:
        return {'copied': False, 'previous_cycle': None}

    carreras_anteriores = list(
        Carrera.objects.filter(ciclo_escolar=ciclo_anterior)
        .prefetch_related('grados__cursos')
        .order_by('id')
    )

    carrera_map = {}
    for carrera_anterior in carreras_anteriores:
        carrera_nueva, _ = Carrera.objects.get_or_create(
            ciclo_escolar=ciclo_nuevo,
            nombre=carrera_anterior.nombre,
            defaults={'activo': carrera_anterior.activo},
        )
        if carrera_nueva.activo != carrera_anterior.activo:
            carrera_nueva.activo = carrera_anterior.activo
            carrera_nueva.save(update_fields=['activo'])
        carrera_map[carrera_anterior.id] = carrera_nueva

    grado_map = {}
    for carrera_anterior in carreras_anteriores:
        carrera_nueva = carrera_map[carrera_anterior.id]
        for grado_anterior in carrera_anterior.grados.all().order_by('id'):
            grado_nuevo, _ = Grado.objects.get_or_create(
                carrera=carrera_nueva,
                nombre=grado_anterior.nombre,
                jornada=grado_anterior.jornada,
                seccion=grado_anterior.seccion,
                defaults={
                    'descripcion': grado_anterior.descripcion,
                    'activo': grado_anterior.activo,
                },
            )
            changed = False
            if grado_nuevo.descripcion != grado_anterior.descripcion:
                grado_nuevo.descripcion = grado_anterior.descripcion
                changed = True
            if grado_nuevo.activo != grado_anterior.activo:
                grado_nuevo.activo = grado_anterior.activo
                changed = True
            if changed:
                grado_nuevo.save(update_fields=['descripcion', 'activo'])
            grado_map[grado_anterior.id] = grado_nuevo

            for curso_anterior in grado_anterior.cursos.all().order_by('id'):
                curso_nuevo, _ = Curso.objects.get_or_create(
                    grado=grado_nuevo,
                    nombre=curso_anterior.nombre,
                    defaults={
                        'descripcion': curso_anterior.descripcion,
                        'activo': curso_anterior.activo,
                    },
                )
                curso_changed = False
                if curso_nuevo.descripcion != curso_anterior.descripcion:
                    curso_nuevo.descripcion = curso_anterior.descripcion
                    curso_changed = True
                if curso_nuevo.activo != curso_anterior.activo:
                    curso_nuevo.activo = curso_anterior.activo
                    curso_changed = True
                if curso_changed:
                    curso_nuevo.save(update_fields=['descripcion', 'activo'])

    return {'copied': True, 'previous_cycle': ciclo_anterior}

def _get_establecimiento(est_id):
    return get_object_or_404(Establecimiento, pk=est_id)


def _get_ciclo(est_id, ciclo_id):
    return get_object_or_404(CicloEscolar.objects.select_related('establecimiento'), pk=ciclo_id, establecimiento_id=est_id)

def _get_carrera(est_id, ciclo_id, car_id):
    return get_object_or_404(
        Carrera.objects.select_related('ciclo_escolar', 'ciclo_escolar__establecimiento'),
        pk=car_id,
        ciclo_escolar_id=ciclo_id,
        ciclo_escolar__establecimiento_id=est_id,
    )

def _get_carrera(est_id, ciclo_id, car_id):
    return get_object_or_404(
        Carrera.objects.select_related('ciclo_escolar', 'ciclo_escolar__establecimiento'),
        pk=car_id,
        ciclo_escolar_id=ciclo_id,
        ciclo_escolar__establecimiento_id=est_id,
    )




def _get_grado(est_id, ciclo_id, car_id, grado_id):
    return get_object_or_404(
        Grado.objects.select_related('carrera', 'carrera__ciclo_escolar', 'carrera__ciclo_escolar__establecimiento'),
        pk=grado_id,
        carrera_id=car_id,
        carrera__ciclo_escolar_id=ciclo_id,
        carrera__ciclo_escolar__establecimiento_id=est_id,
    )


@login_required
@user_passes_test(_can_manage)
def establecimientos_list(request):
    establecimientos = Establecimiento.objects.all()
    establecimientos = filtrar_por_establecimiento_usuario(establecimientos, request.user, 'id')
    return render(request, 'aulapro/establecimientos_list.html', {'establecimientos': establecimientos})


@login_required
@user_passes_test(_can_manage)
def establecimiento_detail(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    establecimiento = _get_establecimiento(est_id)

    if request.method == 'POST':
        _asignar_gestor_a_establecimiento(request, establecimiento)
        return redirect('empleados:establecimiento_detail', est_id=establecimiento.id)

    ciclos = establecimiento.ciclos_escolares.all().order_by('-anio', '-id')
    gestores_asignados = _gestores_qs_para_establecimiento(establecimiento)

    return render(request, 'aulapro/establecimiento_detail.html', {
        'establecimiento': establecimiento,
        'ciclos': ciclos,
        'ciclo_activo': establecimiento.get_ciclo_activo(),
        'gestores_asignados': gestores_asignados,
        'gestores_disponibles': _gestores_disponibles_qs(),
        'puede_gestionar_gestores': es_admin_total(request.user),
    })


@login_required
@user_passes_test(_can_manage)
def establecimiento_update(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    form = EstablecimientoForm(request.POST or None, request.FILES or None, instance=establecimiento)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Establecimiento actualizado correctamente.')
        return redirect('empleados:establecimiento_detail', est_id=establecimiento.id)

    return render(request, 'aulapro/establecimientos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'titulo': 'Editar establecimiento',
    })


@login_required
@user_passes_test(_can_manage)
def ciclos_list(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclos = establecimiento.ciclos_escolares.all().order_by('-anio', '-id')
    return render(request, 'aulapro/ciclos_list.html', {
        'establecimiento': establecimiento,
        'ciclos': ciclos,
    })


@login_required
@user_passes_test(_can_manage)
def ciclo_create(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    form = CicloEscolarForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        ciclo = form.save(commit=False)
        ciclo.establecimiento = establecimiento
        ciclo.activo = bool(form.cleaned_data.get('activo'))

        try:
            with transaction.atomic():
                if ciclo.activo:
                    establecimiento.ciclos_escolares.filter(activo=True).exclude(pk=ciclo.pk).update(activo=False)
                ciclo.save()
                clone_result = _clone_academic_structure_from_previous_cycle(ciclo)
        except Exception:
            messages.error(request, 'No se pudo crear el ciclo escolar. Intente nuevamente.')
        else:
            if clone_result['copied']:
                messages.success(request, 'Ciclo creado correctamente. Se copió la estructura académica del ciclo anterior.')
            else:
                messages.success(request, 'Ciclo creado correctamente. No existía un ciclo anterior para copiar estructura.')
            return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    return render(request, 'aulapro/ciclos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'titulo': 'Nuevo ciclo escolar',
        'ciclo': None,
        'accion': 'Guardar ciclo',
    })


@login_required
@user_passes_test(_can_manage)
def ciclo_update(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)
    form = CicloEscolarForm(request.POST or None, instance=ciclo)

    if request.method == 'POST' and form.is_valid():
        ciclo = form.save(commit=False)
        ciclo.establecimiento = establecimiento
        ciclo.activo = bool(form.cleaned_data.get('activo'))

        with transaction.atomic():
            if ciclo.activo:
                establecimiento.ciclos_escolares.filter(activo=True).exclude(pk=ciclo.pk).update(activo=False)
            ciclo.save()

        messages.success(request, 'Ciclo escolar actualizado correctamente.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    return render(request, 'aulapro/ciclos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'ciclo': ciclo,
        'titulo': 'Editar ciclo escolar',
        'accion': 'Guardar cambios',
    })


@login_required
@user_passes_test(_can_manage)
@require_POST
def ciclo_activar(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)
    try:
        with transaction.atomic():
            establecimiento.ciclos_escolares.update(activo=False)
            ciclo.activo = True
            ciclo.save(update_fields=['activo'])
    except IntegrityError:
        messages.error(request, 'No se pudo activar el ciclo por un conflicto de integridad. Intente nuevamente.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    messages.success(request, f'El ciclo {ciclo.nombre} ahora está activo.')
    return redirect(request.POST.get('next') or 'empleados:ciclos_list', est_id=establecimiento.id)


@login_required
@user_passes_test(_can_manage)
@require_POST
def ciclo_delete(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)

    if ciclo.matriculas.exists():
        messages.warning(request, 'No se puede eliminar el ciclo porque tiene matrículas asociadas.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    if ciclo.activo:
        messages.warning(request, 'No se puede eliminar el ciclo activo. Active otro ciclo primero.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    ciclo.delete()
    messages.success(request, 'Ciclo escolar eliminado correctamente.')
    return redirect('empleados:ciclos_list', est_id=establecimiento.id)



@login_required
@user_passes_test(_can_manage)
def ciclo_detail(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carreras = ciclo.carreras.all().order_by('nombre')
    form = CarreraForm(initial={'ciclo_escolar': ciclo, 'activo': True})
    return render(request, 'aulapro/ciclo_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carreras': carreras,
        'form_carrera': form,
    })




@login_required
@user_passes_test(_can_manage)
def carrera_create(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    ciclo = _get_ciclo(est_id, ciclo_id)
    form = CarreraForm(request.POST or None, initial={'ciclo_escolar': ciclo, 'activo': True})
    if request.method == 'POST' and form.is_valid():
        carrera = form.save(commit=False)
        carrera.ciclo_escolar = ciclo
        carrera.save()
        messages.success(request, 'Carrera creada correctamente.')
        return redirect('empleados:ciclo_detail', est_id=est_id, ciclo_id=ciclo_id)

    return render(request, 'aulapro/carreras/form.html', {
        'establecimiento': ciclo.establecimiento,
        'ciclo': ciclo,
        'form': form,
        'titulo': 'Nueva carrera',
    })

@login_required
@user_passes_test(_can_manage)
def carrera_detail(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grados = Grado.objects.filter(carrera=carrera).order_by('nombre')
    return render(request, 'aulapro/carrera_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grados': grados,
    })


@login_required
@user_passes_test(_can_manage)
def grado_create(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = get_object_or_404(Establecimiento, id=est_id)
    ciclo = get_object_or_404(CicloEscolar, id=ciclo_id, establecimiento=establecimiento)
    carrera = get_object_or_404(Carrera, id=car_id, ciclo_escolar=ciclo)
    grado = None

    form = GradoForm(request.POST or None, initial={'carrera': carrera, 'activo': True})
    if request.method == 'POST' and form.is_valid():
        grado = form.save(commit=False)
        grado.carrera = carrera
        grado.save()
        messages.success(request, 'Grado creado correctamente.')
        return redirect('empleados:carrera_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id)

    return render(request, 'aulapro/grados/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'titulo': f'Nuevo grado - {carrera.nombre}',
        'form': form,
    })



@login_required
def carreras_list(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    return ciclo_detail(request, est_id, ciclo_id)


@login_required
@user_passes_test(_can_manage)
def carrera_update(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    form = CarreraForm(request.POST or None, instance=carrera)
    if request.method == 'POST' and form.is_valid():
        carrera = form.save(commit=False)
        carrera.ciclo_escolar = ciclo
        carrera.save()
        messages.success(request, 'Carrera actualizada correctamente.')
        return redirect('empleados:carrera_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id)

    return render(request, 'aulapro/carreras/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'form': form,
        'titulo': 'Editar carrera',
    })


@login_required
@user_passes_test(_can_manage)
def grados_list(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    return carrera_detail(request, est_id, ciclo_id, car_id)


@login_required
@user_passes_test(_can_manage)
def grado_update(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    form = GradoForm(request.POST or None, instance=grado)
    if request.method == 'POST' and form.is_valid():
        grado = form.save(commit=False)
        grado.carrera = carrera
        grado.save()
        messages.success(request, 'Grado actualizado correctamente.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    return render(request, 'aulapro/grados/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'form': form,
        'titulo': 'Editar grado',
    })




@login_required
@user_passes_test(_can_manage)
def matricula_masiva_grado(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)

    ciclo_escolar = carrera.ciclo_escolar if carrera else None
    if not ciclo_escolar:
        messages.error(request, 'El grado no tiene ciclo escolar asociado. No se puede matricular masivamente.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    establecimiento_usuario = obtener_establecimiento_usuario(request.user)
    if establecimiento_usuario and establecimiento_usuario.id != establecimiento.id:
        messages.error(request, 'No tiene permisos para matricular en este establecimiento.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    session_key = f'matricula_masiva_grado_sel_{grado.id}'
    seleccionados_ids = request.session.get(session_key, [])
    if not isinstance(seleccionados_ids, list):
        seleccionados_ids = []

    def _redirect_self():
        q_param = (request.POST.get('q') or request.GET.get('q') or '').strip()
        if q_param:
            return redirect(f"{request.path}?q={q_param}")
        return redirect(request.path)

    if request.method == 'POST':
        action = (request.POST.get('action') or 'enroll').strip()
        estado = request.POST.get('estado', 'activo')

        if action == 'add':
            alumno_id_raw = request.POST.get('alumno_id', '')
            if not str(alumno_id_raw).isdigit():
                messages.warning(request, 'Alumno inválido para agregar.')
                return _redirect_self()

            alumno_id = int(alumno_id_raw)
            if not Empleado.objects.filter(pk=alumno_id).exists():
                messages.warning(request, 'El alumno seleccionado no existe.')
                return _redirect_self()

            if alumno_id not in seleccionados_ids:
                seleccionados_ids.append(alumno_id)
                request.session[session_key] = seleccionados_ids
                request.session.modified = True
                messages.success(request, 'Alumno agregado a la lista.')
            else:
                messages.info(request, 'El alumno ya estaba en la lista.')
            return _redirect_self()

        if action == 'remove':
            alumno_id_raw = request.POST.get('alumno_id', '')
            if str(alumno_id_raw).isdigit():
                alumno_id = int(alumno_id_raw)
                if alumno_id in seleccionados_ids:
                    seleccionados_ids = [a_id for a_id in seleccionados_ids if a_id != alumno_id]
                    request.session[session_key] = seleccionados_ids
                    request.session.modified = True
                    messages.success(request, 'Alumno removido de la lista.')
            return _redirect_self()

        if action == 'clear':
            request.session[session_key] = []
            request.session.modified = True
            messages.success(request, 'Lista de alumnos seleccionados limpiada.')
            return _redirect_self()

        alumno_ids = list(dict.fromkeys([int(v) for v in seleccionados_ids if str(v).isdigit()]))

        if not alumno_ids:
            messages.warning(request, 'Debe agregar al menos un alumno a la lista.')
            return _redirect_self()

        alumnos_qs = Empleado.objects.filter(id__in=alumno_ids)
        alumnos_map = {a.id: a for a in alumnos_qs}

        inscritos = 0
        omitidos = 0
        errores = 0

        for alumno_id in alumno_ids:
            alumno = alumnos_map.get(alumno_id)
            if not alumno:
                errores += 1
                continue

            if Matricula.objects.filter(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar).exists():
                omitidos += 1
                continue

            matricula = Matricula(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar, estado=estado)
            try:
                matricula.full_clean()
                matricula.save()
                inscritos += 1
            except ValidationError:
                errores += 1

        if inscritos:
            messages.success(request, f'{inscritos} alumnos matriculados correctamente.')
        if omitidos:
            messages.warning(request, f'{omitidos} alumnos ya tenían matrícula y fueron omitidos.')
        if errores:
            messages.error(request, f'{errores} alumnos no pudieron matricularse por validación o permisos.')

        request.session[session_key] = []
        request.session.modified = True

        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    q = (request.GET.get('q') or '').strip()
    alumnos_resultado = Empleado.objects.none()
    if len(q) >= 2:
        alumnos_resultado = (
            Empleado.objects
            .select_related('grado', 'establecimiento')
            .filter(
                Q(codigo_personal__icontains=q)
                | Q(nombres__icontains=q)
                | Q(apellidos__icontains=q)
            )
            .order_by('codigo_personal', 'apellidos', 'nombres')[:25]
        )

    seleccionados_qs = Empleado.objects.select_related('grado', 'establecimiento').filter(id__in=seleccionados_ids)
    seleccionados_map = {a.id: a for a in seleccionados_qs}
    alumnos_seleccionados = [seleccionados_map[a_id] for a_id in seleccionados_ids if a_id in seleccionados_map]

    return render(request, 'aulapro/matricula_masiva_grado.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'ciclo_escolar_contexto': ciclo_escolar,
        'q': q,
        'alumnos_resultado': alumnos_resultado,
        'alumnos_seleccionados': alumnos_seleccionados,
        'estados': Matricula.ESTADOS,
        'estado_default': 'activo',
    })

@login_required
@user_passes_test(_can_manage)
def grado_detail(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)

    ciclo_activo = establecimiento.get_ciclo_activo()
    filtro_form = MatriculaFiltroForm(request.GET or None, establecimiento=establecimiento)
    matriculas = Matricula.objects.select_related('alumno', 'ciclo_escolar').filter(grado=grado)

    ciclo_filtrado = None
    if filtro_form.is_valid():
        estado = filtro_form.cleaned_data.get('estado')
        ciclo_filtrado = filtro_form.cleaned_data.get('ciclo_escolar')
        if estado:
            matriculas = matriculas.filter(estado=estado)

    if ciclo_filtrado:
        matriculas = matriculas.filter(ciclo_escolar=ciclo_filtrado)
    elif ciclo_activo:
        matriculas = matriculas.filter(ciclo_escolar=ciclo_activo)

    configuracion = ConfiguracionGeneral.objects.first()
    layout = establecimiento.get_layout()
    photo_item = (((layout or {}).get("front") or {}).get("items") or {}).get("photo") if isinstance(layout, dict) else {}
    photo_w = int(photo_item.get("w", 250)) if isinstance(photo_item, dict) else 250
    photo_h = int(photo_item.get("h", 350)) if isinstance(photo_item, dict) else 350
    photo_w = photo_w if photo_w > 0 else 250
    photo_h = photo_h if photo_h > 0 else 350
    photo_ratio = photo_w / photo_h
    photo_shape = str(photo_item.get("shape") or "rounded").lower() if isinstance(photo_item, dict) else "rounded"
    if photo_shape not in {"rounded", "square", "circle"}:
        photo_shape = "rounded"
    orientation, canvas_width, canvas_height = resolve_gafete_dimensions(establecimiento, layout)
    layout['canvas'] = {'width': canvas_width, 'height': canvas_height, 'orientation': orientation}
    return render(request, 'aulapro/grado_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'matriculas': matriculas.order_by('-created_at', 'alumno__apellidos'),
        'filtro_form': filtro_form,
        'ciclo_activo': ciclo_activo,
        'configuracion': configuracion,
        'layout': layout,
        'canvas_width': canvas_width,
        'canvas_height': canvas_height,
        'gafete_w': canvas_width,
        'gafete_h': canvas_height,
        'orientacion': orientation,
        'foto_guia_ratio': photo_ratio,
        'foto_guia_shape': photo_shape,
        'foto_guia_w': photo_w,
        'foto_guia_h': photo_h,
    })


def _validate_image_payload(image_bytes):
    if len(image_bytes) > MAX_ALUMNO_IMAGE_SIZE:
        return "La imagen excede el tamaño máximo permitido (5 MB)."
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            img.verify()
    except (UnidentifiedImageError, OSError):
        return "La imagen no es válida o está dañada."
    return None


def _normalize_image_payload(image_bytes, output_ext="jpg"):
    with Image.open(io.BytesIO(image_bytes)) as img:
        normalized = ImageOps.exif_transpose(img)
        out = io.BytesIO()
        ext = str(output_ext or "jpg").lower()
        if ext in {"jpg", "jpeg"}:
            normalized = normalized.convert("RGB")
            normalized.save(out, format="JPEG", quality=92, optimize=True)
            return out.getvalue(), "jpg"
        if ext == "png":
            normalized.save(out, format="PNG", optimize=True)
            return out.getvalue(), "png"
        if ext == "webp":
            normalized.save(out, format="WEBP", quality=92, method=6)
            return out.getvalue(), "webp"
        normalized = normalized.convert("RGB")
        normalized.save(out, format="JPEG", quality=92, optimize=True)
        return out.getvalue(), "jpg"


@login_required
@user_passes_test(_can_manage)
@require_POST
def guardar_foto_alumno_grado(request, est_id, ciclo_id, car_id, grado_id, matricula_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return JsonResponse({"ok": False, "message": "No tiene permisos para ese establecimiento."}, status=403)

    _get_establecimiento(est_id)
    _get_ciclo(est_id, ciclo_id)
    _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    matricula = get_object_or_404(Matricula.objects.select_related("alumno"), pk=matricula_id, grado=grado)
    alumno = matricula.alumno

    uploaded = request.FILES.get("imagen")
    captured_data = (request.POST.get("captured_image") or "").strip()
    content_file = None

    if uploaded:
        ext = os.path.splitext(uploaded.name or "")[1].lower().lstrip(".")
        mime = (uploaded.content_type or "").split(";")[0].lower()
        if mime not in ALUMNO_IMAGE_MIME_EXT and ext not in {"jpg", "jpeg", "png", "webp"}:
            return JsonResponse({"ok": False, "message": "Formato no permitido. Use JPG, PNG o WEBP."}, status=400)
        if uploaded.size > MAX_ALUMNO_IMAGE_SIZE:
            return JsonResponse({"ok": False, "message": "La imagen excede el tamaño máximo permitido (5 MB)."}, status=400)
        image_bytes = uploaded.read()
        payload_error = _validate_image_payload(image_bytes)
        if payload_error:
            return JsonResponse({"ok": False, "message": payload_error}, status=400)
        safe_ext = ALUMNO_IMAGE_MIME_EXT.get(mime, "jpg" if ext == "jpeg" else (ext or "jpg"))
        image_bytes, safe_ext = _normalize_image_payload(image_bytes, safe_ext)
        content_file = ContentFile(image_bytes, name=f"alumno_{alumno.id}_{uuid.uuid4().hex[:8]}.{safe_ext}")
    elif captured_data:
        if not captured_data.startswith("data:image/"):
            return JsonResponse({"ok": False, "message": "Formato de captura inválido."}, status=400)
        try:
            header, encoded = captured_data.split(",", 1)
            mime = header.split(";")[0].replace("data:", "").lower()
            if mime not in ALUMNO_IMAGE_MIME_EXT:
                return JsonResponse({"ok": False, "message": "Formato de captura no permitido."}, status=400)
            image_bytes = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error):
            return JsonResponse({"ok": False, "message": "No se pudo procesar la captura de cámara."}, status=400)
        payload_error = _validate_image_payload(image_bytes)
        if payload_error:
            return JsonResponse({"ok": False, "message": payload_error}, status=400)
        ext = ALUMNO_IMAGE_MIME_EXT[mime]
        image_bytes, ext = _normalize_image_payload(image_bytes, ext)
        content_file = ContentFile(image_bytes, name=f"alumno_{alumno.id}_{uuid.uuid4().hex[:8]}.{ext}")
    else:
        return JsonResponse({"ok": False, "message": "Debe seleccionar un archivo o tomar una foto."}, status=400)

    if alumno.imagen:
        alumno.imagen.delete(save=False)
    alumno.imagen = content_file
    alumno.save(update_fields=["imagen", "updated_at"])

    return JsonResponse({
        "ok": True,
        "message": "Foto del alumno actualizada correctamente.",
        "foto_url": alumno.imagen.url if alumno.imagen else "",
        "alumno_id": alumno.id,
        "matricula_id": matricula.id,
    })


@login_required
@user_passes_test(_can_manage)
def cursos_list(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    cursos = Curso.objects.filter(grado=grado).order_by("nombre")
    return render(request, 'aulapro/grados/cursos/list.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'cursos': cursos,
    })


@login_required
@user_passes_test(_can_manage)
def curso_create(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = None
    form = CursoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        curso = form.save(commit=False)
        curso.grado = grado
        curso.save()
        messages.success(request, 'Curso creado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    return render(request, 'aulapro/grados/cursos/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'titulo': 'Nuevo curso',
        'form': form,
    })


@login_required
@user_passes_test(_can_manage)
def curso_update(request, est_id, ciclo_id, car_id, grado_id, curso_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = get_object_or_404(Curso, pk=curso_id, grado=grado)
    form = CursoForm(request.POST or None, instance=curso)
    if request.method == 'POST' and form.is_valid():
        curso = form.save(commit=False)
        curso.grado = grado
        curso.save()
        messages.success(request, 'Curso actualizado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    return render(request, 'aulapro/grados/cursos/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'titulo': 'Editar curso',
        'form': form,
    })


@login_required
@user_passes_test(_can_manage)
def curso_asignar_docente(request, est_id, ciclo_id, car_id, grado_id, curso_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = get_object_or_404(Curso, pk=curso_id, grado=grado)
    form = AsignarDocenteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(curso)
        messages.success(request, 'Docente asignado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    asignaciones = CursoDocente.objects.filter(curso=curso).select_related('docente').order_by('docente__first_name', 'docente__last_name')
    return render(request, 'aulapro/grados/cursos/asignar_docente.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'form': form,
        'asignaciones': asignaciones,
    })


def _get_docente_cursos_qs(user):
    cursos_qs = (
        CursoDocente.objects.select_related(
            'curso',
            'curso__grado',
            'curso__grado__carrera',
            'curso__grado__carrera__ciclo_escolar',
            'curso__grado__carrera__ciclo_escolar__establecimiento',
            'docente',
        )
        .filter(
            activo=True,
            curso__activo=True,
            curso__grado__carrera__ciclo_escolar__activo=True,
        )
    )

    if _is_docente(user):
        cursos_qs = cursos_qs.filter(docente=user)
    elif es_gestor(user):
        establecimiento = obtener_establecimiento_usuario(user)
        if establecimiento:
            cursos_qs = cursos_qs.filter(curso__grado__carrera__ciclo_escolar__establecimiento=establecimiento)
        else:
            cursos_qs = cursos_qs.none()

    return (
        cursos_qs
        .annotate(
            alumnos_total=Count(
                'curso__grado__matriculas__alumno',
                filter=Q(curso__grado__matriculas__ciclo_escolar=F('curso__grado__carrera__ciclo_escolar')),
                distinct=True,
            ),
            asistencias_total=Count('asistencias', distinct=True),
            detalles_total=Count('asistencias__detalles', distinct=True),
            presentes_total=Count('asistencias__detalles', filter=Q(asistencias__detalles__presente=True), distinct=True),
        )
        .order_by('curso__nombre', 'curso__grado__nombre')
        .distinct()
    )


@login_required
@user_passes_test(_can_manage)
@require_POST
def curso_desasignar_docente(request, est_id, ciclo_id, car_id, grado_id, curso_id, asignacion_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    curso = get_object_or_404(Curso, pk=curso_id, grado_id=grado_id)
    asignacion = get_object_or_404(CursoDocente.objects.select_related('docente'), pk=asignacion_id, curso=curso)

    if not asignacion.activo:
        messages.info(request, 'La asignación ya se encuentra desasignada.')
    else:
        asignacion.activo = False
        asignacion.save(update_fields=['activo'])
        messages.success(request, 'Docente desasignado correctamente.')

    return redirect('empleados:curso_asignar_docente', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id, curso_id=curso_id)


@login_required
@user_passes_test(_is_docente)
def dashboard_docente(request):
    cursos_docente_qs = _get_docente_cursos_qs(request.user)

    cursos_docente = list(cursos_docente_qs)
    curso_docente_ids = [cd.id for cd in cursos_docente]
    grado_ids = {cd.curso.grado_id for cd in cursos_docente if cd.curso_id and cd.curso.grado_id}
    ciclo_ids = {cd.curso.grado.carrera.ciclo_escolar_id for cd in cursos_docente if cd.curso_id and cd.curso.grado_id and cd.curso.grado.carrera_id}

    matriculas_qs = Matricula.objects.filter(grado_id__in=grado_ids, ciclo_escolar_id__in=ciclo_ids)
    total_alumnos_unicos = matriculas_qs.values('alumno_id').distinct().count() if grado_ids else 0

    asistencias_qs = Asistencia.objects.filter(curso_docente_id__in=curso_docente_ids)
    total_asistencias = asistencias_qs.count() if curso_docente_ids else 0
    asistencia_hoy = asistencias_qs.filter(fecha=timezone.localdate()).count() if curso_docente_ids else 0

    detalles_qs = AsistenciaDetalle.objects.filter(asistencia__curso_docente_id__in=curso_docente_ids)
    detalles_totales = detalles_qs.count() if curso_docente_ids else 0
    presentes_totales = detalles_qs.filter(presente=True).count() if curso_docente_ids else 0
    porcentaje_general = round((presentes_totales / detalles_totales) * 100, 2) if detalles_totales else 0

    resumen_cursos = []
    for cd in cursos_docente:
        detalles_curso = getattr(cd, 'detalles_total', 0) or 0
        presentes_curso = getattr(cd, 'presentes_total', 0) or 0
        porcentaje_curso = round((presentes_curso / detalles_curso) * 100, 2) if detalles_curso else 0
        resumen_cursos.append(
            {
                'curso_docente': cd,
                'curso': cd.curso,
                'grado': cd.curso.grado,
                'alumnos_total': getattr(cd, 'alumnos_total', 0) or 0,
                'asistencias_total': getattr(cd, 'asistencias_total', 0) or 0,
                'porcentaje_asistencia': porcentaje_curso,
            }
        )

    ultimas_asistencias = (
        Asistencia.objects.select_related('curso_docente', 'curso_docente__curso', 'curso_docente__curso__grado')
        .filter(curso_docente_id__in=curso_docente_ids)
        .order_by('-fecha', '-id')[:10]
    )

    chart_alumnos_curso_labels = [f"{r['curso'].nombre} ({r['grado'].nombre})" for r in resumen_cursos]
    chart_alumnos_curso_series = [r['alumnos_total'] for r in resumen_cursos]

    asistencias_fecha_rows = (
        asistencias_qs.values('fecha')
        .annotate(total=Count('id'))
        .order_by('-fecha')[:14]
    )
    asistencias_fecha_rows = list(reversed(list(asistencias_fecha_rows)))
    chart_asistencias_fecha_labels = [row['fecha'].strftime('%d/%m') for row in asistencias_fecha_rows]
    chart_asistencias_fecha_series = [row['total'] for row in asistencias_fecha_rows]

    context = {
        'resumen': {
            'cursos_total': len(cursos_docente),
            'alumnos_unicos_total': total_alumnos_unicos,
            'asistencias_total': total_asistencias,
            'asistencia_hoy': asistencia_hoy,
            'porcentaje_general': porcentaje_general,
        },
        'cursos_docente': cursos_docente,
        'resumen_cursos': resumen_cursos,
        'ultimas_asistencias': ultimas_asistencias,
        'chart_alumnos_curso_labels': json.dumps(chart_alumnos_curso_labels),
        'chart_alumnos_curso_series': json.dumps(chart_alumnos_curso_series),
        'chart_asistencias_fecha_labels': json.dumps(chart_asistencias_fecha_labels),
        'chart_asistencias_fecha_series': json.dumps(chart_asistencias_fecha_series),
    }
    return render(request, 'aulapro/dashboard_docente.html', context)


@login_required
@user_passes_test(_can_view_attendance)
def mis_cursos_docente(request):
    cursos_docente = list(_get_docente_cursos_qs(request.user))
    titulo = 'Mis cursos' if _is_docente(request.user) else 'Todos los cursos'
    return render(request, 'aulapro/mis_cursos_docente.html', {
        'cursos_docente': cursos_docente,
        'titulo_cursos': titulo,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_dashboard(request):
    if _is_docente(request.user):
        return mis_cursos_docente(request)
    return redirect('empleados:dahsboard')


@login_required
@user_passes_test(_can_view_attendance)
def docente_curso_detail(request, curso_docente_id):
    curso_docente = get_object_or_404(
        CursoDocente.objects.select_related('curso', 'curso__grado', 'curso__grado__carrera', 'curso__grado__carrera__ciclo_escolar', 'curso__grado__carrera__ciclo_escolar__establecimiento'),
        pk=curso_docente_id,
        activo=True,
        **_attendance_filter_for_user(request.user),
    )
    grado = curso_docente.curso.grado
    alumnos = Empleado.objects.filter(matriculas__grado=grado).distinct().order_by('apellidos', 'nombres')
    return render(request, 'docentes/curso_detail.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'grado': grado,
        'alumnos': alumnos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_home(request, curso_docente_id):
    curso_docente = get_object_or_404(CursoDocente.objects.select_related('curso', 'curso__grado'), pk=curso_docente_id, activo=True, **_attendance_filter_for_user(request.user))
    accion = request.GET.get('generar')
    if accion in {'bimestres', 'semestres'}:
        tipo, total = (PeriodoAcademico.TIPO_BIMESTRE, 4) if accion == 'bimestres' else (PeriodoAcademico.TIPO_SEMESTRE, 2)
        if PeriodoAcademico.objects.filter(curso_docente=curso_docente, tipo=tipo).exists():
            etiqueta = 'bimestre' if tipo == PeriodoAcademico.TIPO_BIMESTRE else 'semestre'
            messages.warning(request, f'Ya existe un {etiqueta} creado. Debe eliminarlo antes de crear uno nuevo.')
            return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente.id)
        for i in range(1, total + 1):
            PeriodoAcademico.objects.create(
                curso_docente=curso_docente,
                tipo=tipo,
                numero=i,
                nombre=f"{tipo.title()} {i}",
                activo=True,
            )
        messages.success(request, 'Periodos generados correctamente.')
        return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente.id)

    periodos = PeriodoAcademico.objects.filter(curso_docente=curso_docente).order_by('tipo', 'numero')
    return render(request, 'docentes/asistencia/home_periodos.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'periodos': periodos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_periodo_detail(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso', 'curso_docente__curso__grado'), pk=periodo_id, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    return render(request, 'docentes/asistencia/periodo_detail.html', {
        'periodo': periodo,
        'curso_docente': periodo.curso_docente,
        'curso': periodo.curso_docente.curso,
    })


@login_required
@user_passes_test(_can_view_attendance)
def tomar_asistencia(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso', 'curso_docente__curso__grado'), pk=periodo_id, activo=True, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    curso_docente = periodo.curso_docente
    fecha = request.POST.get('fecha') or request.GET.get('fecha') or str(timezone.localdate())
    grado = curso_docente.curso.grado
    alumnos = list(Empleado.objects.filter(matriculas__grado=grado).distinct().order_by('apellidos', 'nombres'))

    asistencia, _ = Asistencia.objects.get_or_create(curso_docente=curso_docente, periodo=periodo, fecha=fecha)

    for alumno in alumnos:
        AsistenciaDetalle.objects.get_or_create(asistencia=asistencia, alumno=alumno, defaults={'presente': True})

    detalles = list(AsistenciaDetalle.objects.filter(asistencia=asistencia).select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres'))

    if request.method == 'POST':
        for detalle in detalles:
            detalle.presente = f'presente_{detalle.alumno_id}' in request.POST
        AsistenciaDetalle.objects.bulk_update(detalles, ['presente'])
        messages.success(request, 'Asistencia guardada correctamente.')
        return redirect('empleados:docente_historial_asistencias', periodo_id=periodo.id)

    return render(request, 'docentes/asistencia/tomar_asistencia.html', {
        'periodo': periodo,
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'fecha': fecha,
        'detalles': detalles,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_historial_asistencias(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso'), pk=periodo_id, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    asistencias = Asistencia.objects.filter(periodo=periodo).order_by('-fecha')
    rows = []
    for a in asistencias:
        total = a.detalles.count()
        presentes = a.detalles.filter(presente=True).count()
        ausentes = total - presentes
        rows.append({'asistencia': a, 'total': total, 'presentes': presentes, 'ausentes': ausentes})
    return render(request, 'docentes/asistencia/historial.html', {
        'periodo': periodo,
        'curso_docente': periodo.curso_docente,
        'curso': periodo.curso_docente.curso,
        'rows': rows,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_periodo_historial_excel(request, periodo_id):
    periodo = get_object_or_404(
        PeriodoAcademico.objects.select_related(
            'curso_docente',
            'curso_docente__docente',
            'curso_docente__curso',
            'curso_docente__curso__grado',
            'curso_docente__curso__grado__carrera',
            'curso_docente__curso__grado__carrera__ciclo_escolar',
            'curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento',
        ),
        pk=periodo_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )

    curso_docente = periodo.curso_docente
    curso = curso_docente.curso
    grado = curso.grado

    alumnos = list(
        Empleado.objects.filter(matriculas__grado=grado)
        .distinct()
        .order_by('apellidos', 'nombres')
    )

    resumen_qs = AsistenciaDetalle.objects.filter(asistencia__periodo=periodo).values('alumno_id').annotate(
        asistencias=Sum(Case(When(presente=True, then=1), default=0, output_field=IntegerField())),
        inasistencias=Sum(Case(When(presente=False, then=1), default=0, output_field=IntegerField())),
        total_registros=Count('id'),
    )
    resumen_por_alumno = {
        row['alumno_id']: {
            'asistencias': row['asistencias'] or 0,
            'inasistencias': row['inasistencias'] or 0,
            'total_registros': row['total_registros'] or 0,
        }
        for row in resumen_qs
    }

    total_dias_registrados = Asistencia.objects.filter(periodo=periodo).count()
    establecimiento = '-'
    ciclo = '-'
    if curso.grado and curso.grado.carrera and curso.grado.carrera.ciclo_escolar:
        ciclo_obj = curso.grado.carrera.ciclo_escolar
        ciclo = ciclo_obj.nombre
        establecimiento = ciclo_obj.establecimiento.nombre

    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidado'

    style_title(ws, 1, 'Consolidado de asistencia del período', max_col=7)
    label_font = Font(bold=True)

    encabezado = [
        ('Curso:', curso.nombre),
        ('Docente:', _display_name_for_person(curso_docente.docente)),
        ('Período:', periodo.nombre),
        ('Ciclo escolar:', ciclo),
        ('Establecimiento:', establecimiento),
        ('Fecha de generación:', timezone.localdate().strftime('%d/%m/%Y')),
        ('Total de días con asistencia registrada:', total_dias_registrados),
    ]

    row = 3
    for label, value in encabezado:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ['No.', 'Código / carné', 'Alumno', 'Asistencias', 'Inasistencias', 'Total registros', '% Asistencia']
    style_table_header(ws, row, headers)
    ws.freeze_panes = f'A{row + 1}'

    for idx, alumno in enumerate(alumnos, start=1):
        resumen = resumen_por_alumno.get(alumno.id, {'asistencias': 0, 'inasistencias': 0, 'total_registros': 0})
        total_registros = resumen['total_registros']
        porcentaje = (resumen['asistencias'] / total_registros * 100) if total_registros else 0
        style_table_row(
            ws,
            row + idx,
            [
                idx,
                alumno.codigo_personal or '-',
                f'{alumno.apellidos}, {alumno.nombres}',
                resumen['asistencias'],
                resumen['inasistencias'],
                total_registros,
                f'{porcentaje:.2f}%',
            ],
        )

    autosize_columns(ws)
    return workbook_to_response(wb, f'consolidado_periodo_{periodo.id}')


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_detail(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related('curso_docente', 'curso_docente__docente', 'curso_docente__curso', 'periodo'),
        pk=asistencia_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    detalles = asistencia.detalles.select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres')
    presentes = detalles.filter(presente=True).count()
    ausentes = detalles.count() - presentes
    return render(request, 'docentes/asistencia/detail.html', {
        'asistencia': asistencia,
        'periodo': asistencia.periodo,
        'curso_docente': asistencia.curso_docente,
        'curso': asistencia.curso_docente.curso,
        'detalles': detalles,
        'presentes': presentes,
        'ausentes': ausentes,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_excel(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related(
            'periodo', 'curso_docente', 'curso_docente__docente', 'curso_docente__curso',
            'curso_docente__curso__grado', 'curso_docente__curso__grado__carrera',
            'curso_docente__curso__grado__carrera__ciclo_escolar',
            'curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento'
        ),
        pk=asistencia_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    detalles = list(asistencia.detalles.select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres'))

    curso = asistencia.curso_docente.curso
    docente = asistencia.curso_docente.docente
    periodo = asistencia.periodo.nombre if asistencia.periodo else 'Sin periodo'
    ciclo = '-'
    establecimiento = '-'
    if curso.grado and curso.grado.carrera and curso.grado.carrera.ciclo_escolar:
        ciclo = curso.grado.carrera.ciclo_escolar.nombre
        establecimiento = curso.grado.carrera.ciclo_escolar.establecimiento.nombre

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    style_title(ws, 1, 'Asistencia')

    encabezado = [
        ('Curso:', curso.nombre),
        ('Docente:', _display_name_for_person(docente)),
        ('Fecha:', asistencia.fecha.strftime('%d/%m/%Y')),
        ('Periodo:', periodo),
        ('Ciclo escolar:', ciclo),
        ('Establecimiento:', establecimiento),
    ]

    row = 3
    label_font = Font(bold=True)
    for label, value in encabezado:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ['No.', 'Código / Carné', 'Alumno', 'Estado', 'Observación']
    style_table_header(ws, row, headers)
    ws.freeze_panes = f'A{row + 1}'

    presentes = 0
    ausentes = 0
    for idx, detalle in enumerate(detalles, start=1):
        estado = 'Presente' if detalle.presente else 'Ausente'
        if detalle.presente:
            presentes += 1
        else:
            ausentes += 1
        style_table_row(
            ws,
            row + idx,
            [
                idx,
                detalle.alumno.codigo_personal or '-',
                f'{detalle.alumno.apellidos}, {detalle.alumno.nombres}',
                estado,
                '-',
            ],
        )

    total_row = row + len(detalles) + 2
    ws.cell(row=total_row, column=1, value='Total presentes').font = label_font
    ws.cell(row=total_row, column=2, value=presentes)
    ws.cell(row=total_row + 1, column=1, value='Total ausentes').font = label_font
    ws.cell(row=total_row + 1, column=2, value=ausentes)

    autosize_columns(ws)
    return workbook_to_response(wb, f'asistencia_{asistencia.fecha}')


@login_required
@user_passes_test(_can_view_attendance)
def docente_alumno_historial(request, curso_docente_id, alumno_id):
    curso_docente = get_object_or_404(CursoDocente.objects.select_related('curso', 'curso__grado'), pk=curso_docente_id, activo=True, **_attendance_filter_for_user(request.user))
    alumno = get_object_or_404(Empleado, pk=alumno_id)
    detalles = AsistenciaDetalle.objects.select_related('asistencia', 'asistencia__periodo').filter(
        asistencia__curso_docente=curso_docente,
        alumno=alumno,
    ).order_by('-asistencia__fecha')
    presentes = detalles.filter(presente=True).count()
    ausentes = detalles.count() - presentes

    resumen_periodos = {}
    for d in detalles:
        key = d.asistencia.periodo.nombre if d.asistencia.periodo else 'Sin periodo'
        resumen_periodos.setdefault(key, {'presentes': 0, 'ausentes': 0})
        if d.presente:
            resumen_periodos[key]['presentes'] += 1
        else:
            resumen_periodos[key]['ausentes'] += 1

    return render(request, 'docentes/alumno_historial.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'alumno': alumno,
        'detalles': detalles,
        'presentes': presentes,
        'ausentes': ausentes,
        'resumen_periodos': resumen_periodos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_alumno_historial_excel(request, curso_docente_id, alumno_id):
    curso_docente = get_object_or_404(
        CursoDocente.objects.select_related('curso', 'curso__grado'),
        pk=curso_docente_id,
        activo=True,
        **_attendance_filter_for_user(request.user),
    )
    alumno = get_object_or_404(Empleado, pk=alumno_id)
    detalles = list(
        AsistenciaDetalle.objects.select_related('asistencia', 'asistencia__periodo')
        .filter(asistencia__curso_docente=curso_docente, alumno=alumno)
        .order_by('-asistencia__fecha')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial alumno'

    style_title(ws, 1, 'Historial de asistencia por alumno')
    label_font = Font(bold=True)
    ws.cell(row=3, column=1, value='Curso:').font = label_font
    ws.cell(row=3, column=2, value=curso_docente.curso.nombre)
    ws.cell(row=4, column=1, value='Alumno:').font = label_font
    ws.cell(row=4, column=2, value=f'{alumno.apellidos}, {alumno.nombres}')

    header_row = 6
    style_table_header(ws, header_row, ['No.', 'Fecha', 'Periodo', 'Estado'])
    ws.freeze_panes = f'A{header_row + 1}'

    presentes = 0
    ausentes = 0
    for idx, detalle in enumerate(detalles, start=1):
        estado = 'Presente' if detalle.presente else 'Ausente'
        if detalle.presente:
            presentes += 1
        else:
            ausentes += 1
        style_table_row(
            ws,
            header_row + idx,
            [
                idx,
                detalle.asistencia.fecha.strftime('%d/%m/%Y'),
                detalle.asistencia.periodo.nombre if detalle.asistencia.periodo else 'Sin periodo',
                estado,
            ],
        )

    total_row = header_row + len(detalles) + 2
    ws.cell(row=total_row, column=1, value='Total presentes').font = label_font
    ws.cell(row=total_row, column=2, value=presentes)
    ws.cell(row=total_row + 1, column=1, value='Total ausentes').font = label_font
    ws.cell(row=total_row + 1, column=2, value=ausentes)

    autosize_columns(ws)
    return workbook_to_response(wb, f'historial_alumno_{alumno.id}')




@login_required
@user_passes_test(_can_view_attendance)
@require_POST
def docente_periodo_delete(request, periodo_id):
    periodo = get_object_or_404(
        PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso'),
        pk=periodo_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    curso_docente_id = periodo.curso_docente_id
    total_asistencias = periodo.asistencias.count()
    confirmed = (request.POST.get('confirm_delete') == '1')

    if total_asistencias and not confirmed:
        messages.warning(request, 'Este periodo tiene asistencias asociadas. Confirme la eliminación para continuar.')
        return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente_id)

    if total_asistencias:
        messages.warning(request, f'Se eliminarán {total_asistencias} asistencias asociadas al periodo.')

    periodo.delete()
    messages.success(request, 'Periodo eliminado correctamente.')
    return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente_id)


@login_required
@user_passes_test(_can_manage)
@require_GET
def buscar_alumno(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    _get_grado(est_id, ciclo_id, car_id, grado_id)

    codigo = (request.GET.get('codigo') or '').strip()
    if not codigo:
        return JsonResponse({'found': False, 'error': 'Ingrese un código personal.'}, status=400)

    qs = Empleado.objects.filter(codigo_personal__iexact=codigo)
    if not qs.exists():
        qs = Empleado.objects.filter(codigo_personal__icontains=codigo)

    alumno = qs.order_by('apellidos', 'nombres').first()
    if not alumno:
        return JsonResponse({'found': False, 'error': 'Alumno no encontrado.'}, status=404)

    return JsonResponse({
        'found': True,
        'alumno': {
            'id': alumno.id,
            'codigo': alumno.codigo_personal,
            'nombres': alumno.nombres,
            'apellidos': alumno.apellidos,
            'cui': alumno.cui,
        },
    })


@login_required
@user_passes_test(_can_manage)
@require_POST
def matricular_alumno(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    if not grado.carrera or not grado.carrera.ciclo_escolar_id:
        return JsonResponse({'ok': False, 'error': 'El grado no tiene establecimiento asociado.'}, status=400)

    establecimiento = grado.carrera.ciclo_escolar.establecimiento
    ciclo_activo = establecimiento.get_ciclo_activo()
    if not ciclo_activo:
        return JsonResponse({'ok': False, 'error': 'No hay ciclo escolar activo en este establecimiento. Activa uno para matricular.'}, status=409)

    alumno_id = (request.POST.get('alumno_id') or '').strip()
    if not alumno_id:
        return JsonResponse({'ok': False, 'error': 'Debe seleccionar un alumno.'}, status=400)

    alumno = Empleado.objects.filter(pk=alumno_id).first()
    if not alumno:
        return JsonResponse({'ok': False, 'error': 'Alumno no encontrado.'}, status=404)

    if not ALLOW_MULTI_GRADE_PER_CYCLE:
        other = Matricula.objects.filter(
            alumno=alumno,
            ciclo_escolar=ciclo_activo,
            grado__carrera__ciclo_escolar__establecimiento=establecimiento,
        ).exclude(grado=grado).exists()
        if other:
            return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en otro grado de este establecimiento para el ciclo activo.'}, status=409)

    try:
        matricula, created = Matricula.objects.get_or_create(
            alumno=alumno,
            grado=grado,
            ciclo_escolar=ciclo_activo,
            defaults={
                'estado': 'activo',
                'ciclo': ciclo_activo.anio,
            },
        )
    except IntegrityError:
        return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en este grado y ciclo activo.'}, status=409)

    if not created:
        if matricula.estado != 'activo':
            matricula.estado = 'activo'
            matricula.save(update_fields=['estado'])
        return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en este grado y ciclo activo.'}, status=409)

    return JsonResponse({'ok': True, 'message': 'Alumno matriculado correctamente.'})


@login_required
@user_passes_test(_can_manage)
@require_POST
def desmatricular_alumno(request, matricula_id):
    matricula = get_object_or_404(Matricula, pk=matricula_id)
    matricula.estado = 'inactivo'
    matricula.save(update_fields=['estado'])
    messages.warning(request, 'Matrícula inactivada.')
    return redirect(request.POST.get('next') or 'empleados:establecimientos_list')
